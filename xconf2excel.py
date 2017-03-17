#!/Library/Frameworks/Python.framework/Versions/3.5/bin/python3.5

import logging, sys
import openpyxl, time
import xconf_to_dict
from xconf_to_dict import *
from openpyxl.styles import Font, NamedStyle

logging.basicConfig(filename='logs.txt',level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logging.debug('Start of program')



def xconf2excel(xconfiguration_file):
    """

    Function that takes a VCS backup in the form of a text file with the output of a xconfiguration
    is presented on the web interface of a VCS but using an Excel sheet
    :param xconfiguration_file: File that contains the Output of an xconfiguration from a VCS
    :return: Nothing, but built a Excel sheet with the information contained in a backup, extra information can be seen on log file
    """
    # Build the dictionary that will be used later on to pass the information to the spreadsheet

    dict_user_admin = xconf_to_dict(xconfiguration_file,regex_user_admin)
    dict_group_admin = xconf_to_dict(xconfiguration_file,regex_group_admin)
    dict_sys_admin_services = xconf_to_dict(xconfiguration_file,regex_sys_admin_services)
    dict_discovery_protection = xconf_to_dict(xconfiguration_file,regex_discovery_protection)
    dict_perdomain_dns = xconf_to_dict(xconfiguration_file,regex_perdomain_dns)
    dict_default_dns = xconf_to_dict(xconfiguration_file,regex_default_dns)
    dict_sip_advanced = xconf_to_dict(xconfiguration_file,regex_sip_advanced)
    dict_traversal_media_range = xconf_to_dict(xconfiguration_file,regex_traversal_media_range)
    dict_sip_domain = xconf_to_dict(xconfiguration_file,regex_sip_domain)
    dict_password_security = xconf_to_dict(xconfiguration_file,regex_password_security)
    dict_session_hsts = xconf_to_dict(xconfiguration_file,regex_session_hsts)
    dict_login_source = xconf_to_dict(xconfiguration_file,regex_login_source)
    dict_remote_login = xconf_to_dict(xconfiguration_file,regex_remote_login)
    dict_ip_information = xconf_to_dict(xconfiguration_file,regex_dns_ip)
    dict_system_unit = xconf_to_dict(xconfiguration_file,regex_system_unit)
    dict_ip_protocol = xconf_to_dict(xconfiguration_file,regex_ip_protocol)
    dict_ip_gateway = xconf_to_dict(xconfiguration_file,regex_ip_gateway)
    dict_interface_1_ip = xconf_to_dict(xconfiguration_file,regex_interface_ip_1)
    dict_interface_2_ip = xconf_to_dict(xconfiguration_file,regex_interface_ip_2)
    dict_interface_1_nat = xconf_to_dict(xconfiguration_file,regex_interface_nat_1)
    dict_interface_2_nat = xconf_to_dict(xconfiguration_file,regex_interface_nat_2)
    dict_interface_speed = xconf_to_dict(xconfiguration_file, regex_interface_speed)
    dict_static_routes = xconf_to_dict(xconfiguration_file,regex_static_routes)
    dict_ntp_info = xconf_to_dict(xconfiguration_file,regex_ntp_server)
    dict_snmp = xconf_to_dict(xconfiguration_file,regex_snmp)
    dict_cluster = xconf_to_dict(xconfiguration_file,regex_clustering)
    dict_external_manager = xconf_to_dict(xconfiguration_file,regex_ext_manager)
    dict_h323_mode = xconf_to_dict(xconfiguration_file,regex_h323_mode)
    dict_h323_conf = xconf_to_dict(xconfiguration_file,regex_h323_conf)
    dict_sip_mode = xconf_to_dict(xconfiguration_file,regex_sip_mode)
    dict_sip_conf = xconf_to_dict(xconfiguration_file,regex_sip_info)
    dict_interworking = xconf_to_dict(xconfiguration_file,regex_interworking)
    dict_registration_policy = xconf_to_dict(xconfiguration_file,regex_registration_policy)
    dict_allow_deny_list= xconf_to_dict(xconfiguration_file,regex_registration_allow_deny)
    dict_restriction_policy_servivce= xconf_to_dict(xconfiguration_file,regex_registration_policy_service)
    dict_certificate_authentication = xconf_to_dict(xconfiguration_file,regex_authentication_certificate)
    dict_outbound_credentials = xconf_to_dict(xconfiguration_file,regex_outbound_credentials)
    dict_ads = xconf_to_dict(xconfiguration_file,regex_ads)
    dict_ntlm = xconf_to_dict(xconfiguration_file,regex_ntlm)
    dict_h350 = xconf_to_dict(xconfiguration_file,regex_h350)


  # Create Excel Spreadsheet where the configuration will be presented
    wb = openpyxl.Workbook()
    timestr = time.strftime("%Y%m%d-%H%M%S")
    system_name = dict_system_unit['SystemUnit']['Name']
    destination_filename = system_name+ "_" + timestr + ".xlsx"


    # 1) Obtain System Administrator information (System > Administration)
    ##########################################################################
    ##########################################################################
    sheet = wb.active
    sheet.title = "Sys > Administration"
    styleObj = NamedStyle(name="styleObj")
    styleObj.font= Font(size=16, italic=True) # Using Italic Font and Size 16 for sections title
    wb.add_named_style(styleObj)
    sheet.merge_cells('A1:B1')
    sheet['A1'] = "System name"
    sheet['A1'].style = 'styleObj'
    sheet['A2'] = "System name"
    sheet['B2'] = system_name
    sheet.merge_cells('A3:B3')
    sheet['A3'] = "Ephemeral port range"
    sheet['A3'].style = 'styleObj'
    sheet['A4'] = "Ephemeral port range start"
    sheet['B4'] = dict_ip_information['Ephemeral']['PortRange Start']
    sheet['A5'] = "Ephemeral port range end"
    sheet['B5'] = dict_ip_information['Ephemeral']['PortRange End']
    sheet.merge_cells('A6:B6')
    sheet['A6'] = "Services"
    sheet['A6'].style = 'styleObj'
    sheet['A7'] = "Serial port / console"
    try:
        sheet['B7'] = dict_sys_admin_services['Administration']['SerialConsole']
    except KeyError:
        sheet['B7'] = "Not defined in this version of VCS"
    sheet['A8'] = "SSH service"
    sheet['B8'] = dict_sys_admin_services['Administration']['SSH']
    sheet['A9'] = "Web interface (over HTTPS)"
    sheet['B9'] = dict_sys_admin_services['Administration']['HTTPS']
    sheet.merge_cells('A10:B10')
    sheet['A10'] = "Session limits"
    sheet['A10'].style = 'styleObj'
    sheet['A11'] = "Session time out (minutes)"
    try:
        sheet['B11'] = dict_session_hsts['Management']['Session InactivityTimeout']
    except KeyError:
        sheet['B11'] = "N/A to this version"

    sheet['A12'] = "Per-account session limit"
    try:
        sheet['B12'] = dict_session_hsts['Management']['Session MaxConcurrentSessionsUser']
    except KeyError:
        sheet['B12'] = "N/A to this version"

    sheet['A13'] = "System session limit"
    try:
        sheet['B13'] = dict_session_hsts['Management']['Session MaxConcurrentSessionsTotal']
    except KeyError:
        sheet['B13'] = "N/A to this version"

    sheet.merge_cells('A14:B14')
    sheet['A14'] = "System protection"
    sheet['A14'].style = 'styleObj'
    sheet['A15'] = "Automated protection service"
    sheet['A16'] = "Automatic discovery protection"
    try:
        sheet['B16'] = dict_discovery_protection['DiscoveryProtection']['Mode']
    except KeyError:
        sheet['B16'] = "N/A to this version"
    sheet['A17'] = "Web server configuration"
    sheet['A17'].style = 'styleObj'
    sheet['A18'] = "Redirect HTTP requests to HTTPS"
    sheet['B18'] = "XXXXX Verify with customer XXXXXX"
    sheet['A19'] = "HTTP Strict Transport Security (HSTS)"
    try:
        sheet['B19'] = dict_session_hsts['Management']['Interface HstsMode']
    except KeyError:
        sheet['B19'] = "N/A to this version"

    sheet['A20'] = "Client certificate-based security"
    try:
        sheet['B20'] = dict_certificate_authentication['Authentication']['Certificate']
    except KeyError:
        sheet['B20'] = "N/A to this version"
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 50


    # 2) Obtain Ethernet Details (System > Network Interfaces > Ethernet)
    ##########################################################################
    ##########################################################################
    ip_protocol = dict_ip_protocol['xConfiguration']['IPProtocol']
    ip_gateway = dict_ip_gateway['IP']['Gateway']

    ## Including info into Excel
    wb.create_sheet(title="Sys >Network Interfaces > IP")
    sheet = wb.get_sheet_by_name("Sys >Network Interfaces > IP")

    if dict_interface_2_ip['Ethernet 2']['Address'] == "192.168.0.100":
        sheet.merge_cells('A1:B1')
        sheet['A1'] = "Configuration"
        sheet['A1'].style = 'styleObj'
        sheet['A2'] = "IP protocol"
        sheet['B2'] = ip_protocol
        sheet['A3'] = "IPv4 gateway"
        sheet['B3'] = ip_gateway

        sheet.merge_cells('A4:B4')
        sheet['A4'] = "LAN 1"
        sheet['A4'].style = 'styleObj'
        sheet['A5'] = "IPv4 address"
        sheet['B5'] = dict_interface_1_ip['Ethernet 1']['Address']
        sheet['A6'] = "IPv4 subnet mask"
        sheet['B6'] = dict_interface_1_ip['Ethernet 1']['SubnetMask']
        sheet['A7'] = "Maximum transmission unit (MTU)"
        sheet['B7'] = "XXX Verify value with customer XXXXX"
    else:
        sheet.merge_cells('A1:B1')
        sheet['A1'] = "Configuration"
        sheet['A1'].style = 'styleObj'
        sheet['A2'] = "IP protocol"
        sheet['B2'] = ip_protocol
        sheet['A4'] = "IPv4 gateway"
        sheet['B4'] = ip_gateway
        sheet['A3'] = "External LAN interface"
        sheet['B3'] = dict_ip_gateway['IP']['External Interface']

        sheet.merge_cells('A5:B5')
        sheet['A5'] = "LAN 1"
        sheet['A5'].style = 'styleObj'
        sheet['A6'] = "IPv4 address"
        sheet['B6'] = dict_interface_1_ip['Ethernet 1']['Address']
        sheet['A7'] = "IPv4 subnet mask"
        sheet['B7'] = dict_interface_1_ip['Ethernet 1']['SubnetMask']
        sheet['A8'] = "IPv4 static NAT mode"
        sheet['B8'] = dict_interface_1_nat['Ethernet 1']['StaticNAT Mode']
        sheet['A9'] = "IPv4 static NAT address"
        sheet['B9'] = dict_interface_1_nat['Ethernet 1']['StaticNAT Address']
        sheet['A10'] = "Maximum transmission unit (MTU)"
        sheet['B10'] = "XXX Verify value with customer XXXXX"

        sheet.merge_cells('A11:B11')
        sheet['A11'] = "LAN 2"
        sheet['A11'].style = 'styleObj'
        sheet['A12'] = "IPv4 address"
        sheet['B12'] = dict_interface_2_ip['Ethernet 2']['Address']
        sheet['A13'] = "IPv4 subnet mask"
        sheet['B13'] = dict_interface_2_ip['Ethernet 2']['SubnetMask']
        sheet['A14'] = "IPv4 static NAT mode"
        sheet['B14'] = dict_interface_2_nat['Ethernet 2']['StaticNAT Mode']
        sheet['A15'] = "IPv4 static NAT address"
        sheet['B15'] = dict_interface_2_nat['Ethernet 2']['StaticNAT Address']
        sheet['A16'] = "Maximum transmission unit (MTU)"
        sheet['B16'] = "XXX Verify value with customer XXXXX"

    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 50


    # 3) System > Network Interface > Static Routes
    ##########################################################################
    ##########################################################################
    wb.create_sheet(title="Sys>Net_Int>Static_routes")
    sheet = wb.get_sheet_by_name("Sys>Net_Int>Static_routes")
    if dict_static_routes:
        sheet['A1'] = "IP address"
        sheet['A1'].style = 'styleObj'
        sheet['B1'] = "Prefix Length"
        sheet['B1'].style = 'styleObj'
        sheet['C1'] = "Gateway"
        sheet['C1'].style = 'styleObj'
        sheet['D1'] = "Interface"
        sheet['D1'].style = 'styleObj'

        for e,keys in enumerate(sorted(dict_static_routes),2):
            sheet['A{0}'.format(e)] = ("").join(dict_static_routes[keys]['Address'].split('"'))
            sheet['B{0}'.format(e)] = ("").join(dict_static_routes[keys]['PrefixLength'].split('"'))
            sheet['C{0}'.format(e)] = ("").join(dict_static_routes[keys]['Gateway'].split('"'))
            sheet['D{0}'.format(e)] = ("").join(dict_static_routes[keys]['Interface'].split('"'))

    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 25


    # 3) System > DNS
    ##########################################################################
    ##########################################################################
    wb.create_sheet(title="System > DNS")
    sheet = wb.get_sheet_by_name("System > DNS")
    sheet.merge_cells('A1:D1')
    sheet['A1'] = "DNS Settings"
    sheet['A1'].style = 'styleObj'
    sheet['A2'] = "System host name"
    sheet['A3'] = "Domain name"
    sheet['A4'] = "DNS requets port range"
    sheet['B2'] = dict_ip_information['DNS']['Hostname']
    sheet['B3'] = dict_ip_information['DNS']['Domain Name']

    try:
        if dict_ip_information['DNS']['UseEphemeralPortRange'] == 'Off':
            sheet['B4'] = "Use a custom port range"
            sheet['A5'] = "DNS request port range start"
            sheet['A6'] = "DNS request port range end"
            sheet['B5'] = dict_ip_information['Ephemeral']['PortRange Start']
            sheet['B6'] = dict_ip_information['Ephemeral']['PortRange End']
            sheet.merge_cells('A7:D7')
            sheet['A7'] = "Default DNS servers"
            sheet['A7'].style = 'styleObj'

            for enum, key in enumerate(sorted(dict_default_dns),8):
                sheet['A{0}'.format(enum)].value = "Address {}".format(dict_default_dns[key]['Index'])
                sheet['B{0}'.format(enum)].value = "".join(dict_default_dns[key]['Address'].split('"'))

            sheet.merge_cells('A13:D13')
            sheet['A13'].style = 'styleObj'
            sheet['A13'] = "Per-Domain DNS"

            for enum, key in enumerate(sorted(dict_perdomain_dns),14):
                try:
                    sheet['A{0}'.format(enum)].value = "Address {}".format(key[-1])
                except KeyError:
                    sheet['A{0}'.format(enum)].value = "Address 1"
                try:
                    sheet['B{0}'.format(enum)].value = dict_perdomain_dns[key]['Address']
                except KeyError:
                    sheet['B{0}'.format(enum)].value = ""
                sheet['C{0}'.format(enum)].value = "Domain names"
                try:
                    sheet['D{0}'.format(enum)].value = dict_perdomain_dns[key]['Domain1']
                except KeyError:
                    sheet['D{0}'.format(enum)].value = ""
                try:
                    sheet['E{0}'.format(enum)].value = dict_perdomain_dns[key]['Domain2']
                except KeyError:
                    sheet['E{0}'.format(enum)].value = ""
        else:
            sheet['B4'] = "Use the ephemeral port range"
            sheet.merge_cells('A5:D5')
            sheet['A5'] = "Default DNS servers"
            sheet['A5'].style = 'styleObj'

            for enum, key in enumerate(sorted(dict_default_dns),6):
                sheet['A{0}'.format(enum)].value = "Address {}".format(dict_default_dns[key]['Index'])
                sheet['B{0}'.format(enum)].value = "".join(dict_default_dns[key]['Address'].split('"'))

            sheet.merge_cells('A11:D11')
            sheet['A11'].style = 'styleObj'
            sheet['A11'] = "Per-Domain DNS"

            for enum, key in enumerate(sorted(dict_perdomain_dns),12):
                try:
                    sheet['A{0}'.format(enum)].value = "Address {}".format(key[-1])
                except KeyError:
                    sheet['A{0}'.format(enum)].value = "Address 1"
                try:
                    sheet['B{0}'.format(enum)].value = dict_perdomain_dns[key]['Address']
                except KeyError:
                    sheet['B{0}'.format(enum)].value = ""
                sheet['C{0}'.format(enum)].value = "Domain names"
                try:
                    sheet['D{0}'.format(enum)].value = dict_perdomain_dns[key]['Domain1']
                except KeyError:
                    sheet['D{0}'.format(enum)].value = ""
                try:
                    sheet['E{0}'.format(enum)].value = dict_perdomain_dns[key]['Domain2']
                except KeyError:
                    sheet['E{0}'.format(enum)].value = ""
    except KeyError:
        pass

    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 25
    sheet.column_dimensions['E'].width = 25

    # 4) System > Time
    ##########################################################################
    ##########################################################################
    wb.create_sheet(title="System > Time")
    sheet = wb.get_sheet_by_name("System > Time")
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 25

    sheet.merge_cells('A1:D1')
    sheet['A1'] = "NTP servers"
    sheet['A1'].style = 'styleObj'

    try:
        sheet.merge_cells('A8:D8')
        sheet['A8'] = "Time zone"
        sheet['A8'].style = 'styleObj'
        sheet['A9'] = "Time zone"
        sheet['B9'] = dict_ntp_info['TimeZone']['Name']

        for enum, keys in enumerate(sorted(dict_ntp_info['NTP Server'])):
            sheet['A{0}'.format(enum + 2)].value = "NTP server {}".format(enum + 1)
            sheet['B{0}'.format(enum + 2)].value = "Address"
            sheet['C{0}'.format(enum + 2)].value = dict_ntp_info['NTP Server'][keys]
            sheet['D{0}'.format(enum + 2)].value = "Authentication"
            sheet['E{0}'.format(enum + 2)].value = "XXXX Check with customer XXX"
    except KeyError:
        pass

    # 5) System > SNMP
    ##########################################################################
    ##########################################################################
    wb.create_sheet(title="System > SNMP")
    sheet = wb.get_sheet_by_name("System > SNMP")
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 50

    sheet.merge_cells('A1:B1')
    sheet['A1'] = "Configuration"
    sheet['A1'].style = 'styleObj'

    try:
        if dict_snmp['SNMP']['V2cMode'] == "Off" and dict_snmp['SNMP']['V3Mode'] == "Off":
            sheet['A2'] = "SNMP Mode"
            sheet['B2'] = "Disabled"
        elif dict_snmp['SNMP']['V2cMode'] == "On" and dict_snmp['SNMP']['V3Mode'] == "Off":
            sheet['A2'] = "SNMP Mode"
            sheet['B2'] = "v2c"
            sheet['A3'] = "Description"
            try:
                sheet['B3'] = dict_snmp['SNMP']['SystemDescription']
            except KeyError:
                sheet['B3'] = "Not defined on this VCS Version"
            sheet['A4'] = "Community name"
            sheet['B4'] = dict_snmp['SNMP']['CommunityName']
            sheet['A5'] = "System contact"
            sheet['B5'] = dict_snmp['SNMP']['SystemContact']
            sheet['A6'] = "Location"
            sheet['B6'] = dict_snmp['SNMP']['SystemLocation']
        elif dict_snmp['SNMP']['V2cMode'] == "Off" and dict_snmp['SNMP']['V3Mode'] == "On":
            sheet['A2'] = "SNMP Mode"
            sheet['B2'] = "v3 secure SNMP"
            sheet['A3'] = "Description"
            sheet['B3'] = dict_snmp['SNMP']['SystemDescription']
            sheet['A4'] = "System contact"
            sheet['B4'] = dict_snmp['SNMP']['SystemContact']
            sheet['A5'] = "Location"
            sheet['B5'] = dict_snmp['SNMP']['SystemLocation']
            sheet['A6'] = "Username"
            sheet['B6'] = dict_snmp['SNMP']['V3UserName']
            sheet.merge_cells('A7:B7')
            sheet['A7'] = "v3 Authentication"
            sheet['A7'].style = 'styleObj'
            sheet['A8'] = "Authentication mode"
            sheet['B8'] = dict_snmp['SNMP']['V3AuthenticationMode']
            sheet['A9'] = "Type"
            sheet['B9'] = dict_snmp['SNMP']['V3AuthenticationType']
            sheet['A10'] = "Password"
            sheet['B10'] = dict_snmp['SNMP']['V3AuthenticationPassword']
            sheet.merge_cells('A11:B11')
            sheet['A11'] = "v3 Privacy"
            sheet['A11'].style = 'styleObj'
            sheet['A12'] = "Privacy mode"
            sheet['B12'] = dict_snmp['SNMP']['V3PrivacyMode']
            sheet['A13'] = "Type"
            sheet['B13'] = dict_snmp['SNMP']['V3PrivacyType']
            sheet['A14'] = "Password"
            sheet['B14'] = dict_snmp['SNMP']['V3PrivacyPassword']
        else:
            sheet['A2'] = "SNMP Mode"
            sheet['B2'] = "v3 plus TMS support"
            sheet['A3'] = "Description"
            try:
                sheet['B3'] = dict_snmp['SNMP']['SystemDescription']
            except KeyError:
                sheet['B3'] = "Not defined on this version of VCS"
            sheet['A4'] = "Community name"
            sheet['B4'] = dict_snmp['SNMP']['CommunityName']
            sheet['A5'] = "System contact"
            sheet['B5'] = dict_snmp['SNMP']['SystemContact']
            sheet['A6'] = "Location"
            sheet['B6'] = dict_snmp['SNMP']['SystemLocation']
            sheet['A7'] = "Username"
            sheet['B7'] = dict_snmp['SNMP']['V3UserName']
            sheet.merge_cells('A8:B8')
            sheet['A8'] = "v3 Authentication"
            sheet['A8'].style = 'styleObj'
            sheet['A9'] = "Authentication mode"
            sheet['B9'] = dict_snmp['SNMP']['V3AuthenticationMode']
            sheet['A10'] = "Type"
            sheet['B10'] = dict_snmp['SNMP']['V3AuthenticationType']
            sheet['A11'] = "Password"
            sheet['B11'] = dict_snmp['SNMP']['V3AuthenticationPassword']
            sheet.merge_cells('A12:B12')
            sheet['A12'] = "v3 Privacy"
            sheet['A12'].style = 'styleObj'
            sheet['A13'] = "Privacy mode"
            sheet['B13'] = dict_snmp['SNMP']['V3PrivacyMode']
            sheet['A14'] = "Type"
            sheet['B14'] = dict_snmp['SNMP']['V3PrivacyType']
            sheet['A15'] = "Password"
            sheet['B15'] = dict_snmp['SNMP']['V3PrivacyPassword']
    except KeyError:
        pass

    # 6) System > Clustering
    ##########################################################################
    ##########################################################################

    wb.create_sheet(title="System > Clustering")
    sheet = wb.get_sheet_by_name("System > Clustering")
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 50
    sheet.merge_cells('A1:B1')
    sheet['A1'] = "Configuration"
    sheet['A1'].style = 'styleObj'
    sheet['A2'] = "Cluster name (FQDN for Provisioning)"
    sheet['B2'] = dict_cluster['Alternates']['Cluster Name']
    sheet['A3'] = "Configuration master"
    sheet['B3'] = dict_cluster['Alternates']['ConfigurationMaster']
    sheet['A4'] = "Cluster pre-shared key"
    sheet['B4'] = "******"

    for enum,i in enumerate (range(6)):
        sheet['A{0}'.format(enum + 5)].value = 'Peer {} IP address'.format(enum + 1)
        sheet['B{0}'.format(enum + 5)].value = dict_cluster['Alternates']['Peer {0}'.format(enum + 1)]


    # 6) System > QoS
    ##########################################################################
    ##########################################################################
    wb.create_sheet(title="System > QoS")
    sheet = wb.get_sheet_by_name("System > QoS")
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 50
    sheet.merge_cells('A1:B1')
    sheet['A1'] = "Tagging"
    sheet['A1'].style = 'styleObj'
    try:
        sheet['A2'] = "QoS mode"
        sheet['B2'] = dict_ip_information['QoS']['Mode']
        sheet['A3'] = "Tag value"
        sheet['B3'] = dict_ip_information['QoS']['Value']
    except KeyError:
        dict_qos = xconf_to_dict(xconfiguration_file,regex_qos)
        sheet['A2'] = "DSCP Signaling value"
        sheet['B2'] = dict_qos['QoS']['Signaling']
        sheet['A3'] = "DSCP Audio value"
        sheet['B3'] = dict_qos['QoS']['Audio']
        sheet['A4'] = "DSCP Video value"
        sheet['B4'] = dict_qos['QoS']['Video']
        sheet['A5'] = "DSCP XMPP value"
        sheet['B5'] = dict_qos['QoS']['XMPP']

    # 7) System > External Manager
    ##########################################################################
    ##########################################################################
    wb.create_sheet(title="System > External Manager")
    sheet = wb.get_sheet_by_name("System > External Manager")
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 50
    sheet.merge_cells('A1:B1')
    sheet['A1'] = "Configuration"
    sheet['A1'].style = 'styleObj'
    sheet['A2'] = "Address"
    sheet['B2'] = dict_external_manager['ExternalManager']['Address']
    sheet['A3'] = "Path"
    sheet['B3'] = dict_external_manager['ExternalManager']['Path']
    sheet['A4'] = "Protocol"
    sheet['B4'] = dict_external_manager['ExternalManager']['Protocol']
    sheet['A5'] = "Certificate verification mode"
    sheet['B5'] = dict_external_manager['ExternalManager']['Server Certificate Verification Mode']


    # 8) Configuration > Protocol > H.323
    ##########################################################################
    ##########################################################################
    wb.create_sheet(title="Config > Protocols > H.323")
    sheet = wb.get_sheet_by_name("Config > Protocols > H.323")
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 50
    sheet.merge_cells('A1:B1')
    sheet['A1'] = "Configuration"
    sheet['A1'].style = 'styleObj'
    sheet['A2'] = "H.323 mode"
    sheet['B2'] = dict_h323_mode['H323']['Mode']
    sheet.merge_cells('A3:B3')
    sheet['A3'] = "Gatekeeper"
    sheet['A3'].style = 'styleObj'
    sheet['A4'] = "Registration UDP port"
    sheet['B4'] = dict_h323_conf['Gatekeeper']['Registration UDP Port']
    sheet['A5'] = "Registration conflict mode"
    sheet['B5'] = dict_h323_conf['Gatekeeper']['Registration ConflictMode']
    sheet['A6'] = "Call signalling TCP port"
    sheet['B6'] = dict_h323_conf['Gatekeeper']['CallSignaling TCP Port']
    sheet['A7'] = "Call signalling port range start"
    sheet['B7'] = dict_h323_conf['Gatekeeper']['CallSignaling PortRange Start']
    sheet['A8'] = "Call signalling port range end"
    sheet['B8'] = dict_h323_conf['Gatekeeper']['CallSignaling PortRange End']
    sheet['A9'] = "Time to live"
    sheet['B9'] = dict_h323_conf['Gatekeeper']['TimeToLive']
    sheet['A10'] = "Call time to live"
    sheet['B10'] =dict_h323_conf['Gatekeeper']['CallTimeToLive']
    sheet['A11'] = "Auto discover"
    sheet['B11'] = dict_h323_conf['Gatekeeper']['AutoDiscovery Mode']
    sheet.merge_cells('A12:B12')
    sheet['A12'] = "Gateway"
    sheet['A12'].style = 'styleObj'
    sheet['A13'] = "Caller ID"
    sheet['B13'] = dict_h323_conf['Gateway']['CallerId']


    # 9) Configuration > Protocol > SIP
    ##########################################################################
    ##########################################################################
    wb.create_sheet(title="Config > Protocols > SIP")
    sheet = wb.get_sheet_by_name("Config > Protocols > SIP")
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 50
    sheet.merge_cells('A1:B1')
    sheet['A1'] = "Configuration"
    sheet['A1'].style = 'styleObj'
    sheet['A2'] = "SIP mode"
    sheet['B2'] = dict_sip_mode['SIP']['Mode']
    sheet['A3'] = "UDP mode"
    sheet['B3'] = dict_sip_conf['UDP']['Mode']
    sheet['A4'] = "UDP port"
    sheet['B4'] = dict_sip_conf['UDP']['Port']
    sheet['A5'] = "TCP mode"
    sheet['B5'] = dict_sip_conf['TCP']['Mode']
    sheet['A5'] = "TCP port"
    sheet['B5'] = dict_sip_conf['TCP']['Port']
    sheet['A6'] = "TLS mode"
    sheet['B6'] = dict_sip_conf['TLS']['Mode']
    sheet['A7'] = "TLS port"
    sheet['B7'] = dict_sip_conf['TLS']['Port']
    sheet['A8'] = "MTLS mode"
    sheet['A9'] = "MTLS port"
    try:
        sheet['B8'] = dict_sip_conf['MTLS']['Mode']
        sheet['B9'] = dict_sip_conf['MTLS']['Port']
    except:
        sheet['B8'] = "N/A"
        sheet['B8'] = "N/A"
    sheet['A10'] = "TCP outbound port start"
    sheet['B10'] = dict_sip_conf['TCP']['Outbound Port Start']
    sheet['A11'] = "TCP outbound port end"
    sheet['B11'] = dict_sip_conf['TCP']['Outbound Port End']
    sheet['A12'] = "Session refresh interval (seconds)"
    sheet['B12'] = dict_sip_conf['Session']['Refresh Value']
    sheet['A13'] = "Minimum Session refresh interval (seconds)"
    sheet['B13'] = dict_sip_conf['Session']['Refresh Minimum']
    cert_revocation_chk = dict_sip_conf['TLS']['Certificate Revocation Checking Mode']
    sheet.merge_cells('A14:B14')
    sheet['A14'] = "Certificate revocation checking"
    sheet['A14'].style = 'styleObj'
    sheet['A15'] = "Certificate revocation checking mode"
    sheet['B15'] = cert_revocation_chk

    if cert_revocation_chk == "Off":
        sheet.merge_cells('A16:B16')
        sheet['A16'] = "Registration controls"
        sheet['A16'].style = 'styleObj'
        sheet['A17'] = "Standard registration refresh strategy"
        sheet['B17'] = dict_sip_conf['Registration']['Standard Refresh Strategy']
        sheet['A18'] = "Standard registration refresh minimum (seconds)"
        sheet['B18'] = dict_sip_conf['Registration']['Standard Refresh Minimum']
        sheet['A19'] = "Standard registration refresh maximum (seconds)"
        sheet['B19'] = dict_sip_conf['Registration']['Standard Refresh Maximum']
        sheet['A20'] = "Outbound registration refresh strategy"
        sheet['B20'] = dict_sip_conf['Registration']['Outbound Refresh Strategy']
        sheet['A21'] = "Outbound registration refresh minimum (seconds)"
        sheet['B21'] = dict_sip_conf['Registration']['Outbound Refresh Minimum']
        sheet['A22'] = "Outbound registration refresh maximum (seconds)"
        sheet['B22'] = dict_sip_conf['Registration']['Outbound Refresh Maximum']
        sheet['A23'] = "SIP registration proxy mode"
        sheet['B23'] = dict_sip_conf['Registration']['Proxy Mode']
        sheet.merge_cells('A24:B24')
        sheet['A24'] = "Authentication"
        sheet['A24'].style = 'styleObj'
        sheet['A25'] = "Delegated credential checking"
        sheet['B25'] = "XXX Check with Customer XXXX"
        sheet.merge_cells('A26:B26')
        try:
            sheet['A26'] = "Advanced"
            sheet['A26'].style = 'styleObj'
            sheet['A27'] = "SDP max size"
            sheet['A28'] = "SIP TCP connect timeout"
            sheet['B27'] = dict_sip_advanced['SIP Advanced']['SdpMaxSize']
            sheet['B28'] = dict_sip_advanced['SIP Advanced']['SipTcpConnectTimeout']
        except KeyError:
            logging.debug ("Advanced SIP not defined on this version of VCS")

    elif cert_revocation_chk == "On":
        sheet['A16'] = "Use OCSP"
        sheet['B16'] = dict_sip_conf['TLS']['Certificate Revocation Checking OCSP Mode']
        sheet['A17'] = "Use CRLs"
        sheet['B17'] = dict_sip_conf['TLS']['Revocation Checking CRL Mode']
        sheet['A18'] = "Allow CRL downloads from CDPs"
        sheet['B18'] = dict_sip_conf['TLS']['Certificate Revocation Checking CRL Network Fetch Mode']
        sheet['A19'] = "Fallback behaviour"
        fallbackup = dict_sip_conf['TLS']['Certificate Revocation Checking Source Inaccessibility Behavior']

        if fallbackup == "Fail":
            fallbackup = "Treat as revocked"
        else:
            fallbackup = "Treat as not revocked"

        sheet['B19'] = fallbackup
        sheet.merge_cells('A20:B20')
        sheet['A20'] = "Registration controls"
        sheet['A20'].style = 'styleObj'
        sheet['A21'] = "Standard registration refresh strategy"
        sheet['B21'] = dict_sip_conf['Registration']['Standard Refresh Strategy']
        sheet['A22'] = "Standard registration refresh minimum (seconds)"
        sheet['B22'] = dict_sip_conf['Registration']['Standard Refresh Minimum']
        sheet['A23'] = "Standard registration refresh maximum (seconds)"
        sheet['B23'] = dict_sip_conf['Registration']['Standard Refresh Maximum']
        sheet['A24'] = "Outbound registration refresh strategy"
        sheet['B24'] = dict_sip_conf['Registration']['Outbound Refresh Strategy']
        sheet['A25'] = "Outbound registration refresh minimum (seconds)"
        sheet['B25'] = dict_sip_conf['Registration']['Outbound Refresh Minimum']
        sheet['A26'] = "Outbound registration refresh maximum (seconds)"
        sheet['B26'] = dict_sip_conf['Registration']['Outbound Refresh Maximum']
        sheet['A27'] = "SIP registration proxy mode"
        sheet['B27'] = dict_sip_conf['Registration']['Proxy Mode']
        sheet.merge_cells('A28:B28')
        sheet['A28'] = "Authentication"
        sheet['A28'].style = 'styleObj'
        sheet['A29'] = "Delegated credential checking"
        sheet['B29'] = "XXX Check with customer XXX"
        try:
            sheet.merge_cells('A30:B30')
            sheet['A30'] = "Advanced"
            sheet['A30'].style = 'styleObj'
            sheet['A31'] = "SDP max size"
            sheet['A32'] = "SIP TCP connect timeout"
            sheet['B31'] = dict_sip_advanced['SIP Advanced']['SdpMaxSize']
            sheet['B32'] = dict_sip_advanced['SIP Advanced']['SipTcpConnectTimeout']
        except KeyError:
            logging.debug ("Advanced SIP not defined")


    # 10) Configuration > Protocol > Interworking
    ##########################################################################
    ##########################################################################
    wb.create_sheet(title="Config > Prot > Interworking")
    sheet = wb.get_sheet_by_name("Config > Prot > Interworking")
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 50
    sheet.merge_cells('A1:B1')
    sheet['A1'] = "Configuration"
    sheet['A1'].style = 'styleObj'
    sheet['A2'] = "H.323 <-> SIP interworking mode"
    sheet['B2'] = dict_interworking['Interworking']['Mode']


    # 11) Configuration > Registration > Configuration
    ##########################################################################
    ##########################################################################
    wb.create_sheet(title="Config > Registration")
    sheet = wb.get_sheet_by_name("Config > Registration")
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 25

    sheet.merge_cells('A1:B1')
    sheet['A1'] = "Configuration"
    sheet['A1'].style = 'styleObj'
    sheet['A2'] = "Restriction Policy"
    restriction_policy = dict_registration_policy['RestrictionPolicy']['Mode']
    sheet['B2'] = restriction_policy
    if restriction_policy == "AllowList":
        sheet['A4'] = "Description"
        sheet['B4'] = "Pattern Type"
        sheet['C4'] = "Pattern string"
        sheet.merge_cells('A3:C3')
        sheet['A3'] = "Registration Allow List"
        sheet['A3'].style = 'styleObj'
        for enum, keys in enumerate(sorted(dict_allow_deny_list),5):
            if keys.startswith("Allow"):
                sheet['A{}'.format(enum)].value = dict_allow_deny_list[keys]['Description']
                sheet['B{}'.format(enum)].value = dict_allow_deny_list[keys]['Pattern Type']
                sheet['C{}'.format(enum)].value = dict_allow_deny_list[keys]['Pattern String']
    elif restriction_policy == "DenyList":
        sheet['A4'] = "Description"
        sheet['B4'] = "Pattern Type"
        sheet['C4'] = "Pattern string"
        sheet.merge_cells('A3:C3')
        sheet['A3'] = "Registration Deny List"
        sheet['A3'].style = 'styleObj'
        for enum, keys in enumerate(sorted(dict_allow_deny_list),5):
            if keys.startswith("Deny"):
                sheet['A{}'.format(enum)].value = dict_allow_deny_list[keys]['Description']
                sheet['B{}'.format(enum)].value = dict_allow_deny_list[keys]['Pattern Type']
                sheet['C{}'.format(enum)].value = dict_allow_deny_list[keys]['Pattern String']
    elif restriction_policy == "PolicyService":
        sheet['A3'] = "Protocol"
        sheet['B3'] = dict_restriction_policy_servivce['Service 1']['Protocol']
        sheet['A4'] = "Certificate verification mode"
        sheet['B4'] = dict_restriction_policy_servivce['Service 1']['TLS Verify Mode']
        sheet['A5'] = "HTTPS certificate revocation list (CRL) checking"
        sheet['B5'] = dict_restriction_policy_servivce['Service 1']['TLS CRLCheck Mode']
        sheet['A6'] = "Server 1 address"
        sheet['B6'] = dict_restriction_policy_servivce['Service 1']['Server 1 Address']
        sheet['A7'] = "Server 2 address"
        sheet['B7'] = dict_restriction_policy_servivce['Service 1']['Server 2 Address']
        sheet['A8'] = "Server 3 address"
        sheet['B8'] = dict_restriction_policy_servivce['Service 1']['Server 3 Address']
        sheet['A9'] = "Path"
        sheet['B9'] = dict_restriction_policy_servivce['Service 1']['Path']
        sheet['A10'] = "Status path"
        sheet['B10'] = dict_restriction_policy_servivce['Service 1']['Status Path']
        sheet['A11'] = "Username"
        sheet['B11'] = dict_restriction_policy_servivce['Service 1']['UserName']
        sheet['A12'] = "Password"
        sheet['B12'] = dict_restriction_policy_servivce['Service 1']['Password']
        sheet['A13'] = "Default CPL"
        sheet['B13'] = dict_restriction_policy_servivce['Service 1']['DefaultCPL']


    # 11) Configuration > Authentication
    ##########################################################################
    ##########################################################################
    # a) Outbound credentials
    wb.create_sheet(title="Config > Auth > Outbound")
    sheet = wb.get_sheet_by_name("Config > Auth > Outbound")
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 25

    sheet.merge_cells('A1:D1')
    sheet['A1'] = "Configuration"
    sheet['A1'].style = 'styleObj'
    sheet['A2'] = "Authentication username"
    sheet['B2'] = dict_outbound_credentials['Authentication']['UserName']
    sheet['A3'] = "Authentication password"
    try:
        sheet['B3'] = dict_outbound_credentials["Authentication"]['Password']
    except KeyError:
        sheet['B3'] = "******"

    # b) Local Database
    wb.create_sheet(title="Config > Auth > Devices > db")
    sheet = wb.get_sheet_by_name("Config > Auth > Devices > db")
    sheet.column_dimensions['A'].width = 60
    sheet.column_dimensions['B'].width = 60
    sheet.column_dimensions['C'].width = 60

    dict_credentials = xconf_to_dict(xconfiguration_file,regex_localDB)

    sheet.merge_cells('A1:D1')
    sheet['A1'] = "Local authentication database"
    sheet['A1'].style = 'styleObj'
    sheet['A2'] = "Name"
    sheet['B2'] = "Password"

    for enum, keys in enumerate(dict_credentials,3):
        sheet["A{}".format(enum)]= dict_credentials[keys]['Name']
        sheet["B{}".format(enum)]= dict_credentials[keys]['Password']

    # c) Active Directory Services
    wb.create_sheet(title="Config > Auth > Devices > AD")
    sheet = wb.get_sheet_by_name("Config > Auth > Devices > AD")
    sheet.column_dimensions['A'].width = 60
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 25

    sheet.merge_cells('A1:D1')
    sheet['A1'] = "Configuration"
    sheet['A1'].style = 'styleObj'
    auth_ad_connect = dict_ads['ADS']['Mode']
    sheet['A2'] = "Connect to Active Directory Service"
    sheet['B2'] = auth_ad_connect
    sheet['A3'] = "NTLM protocol challenges"
    sheet['B3'] = dict_ntlm['NTLM']['Mode']

    if auth_ad_connect == "On":
        sheet.merge_cells('A4:D4')
        sheet['A4'] = "Active Directory configuration"
        sheet['A4'].style = 'styleObj'
        sheet['A5'] = "AD domain"
        sheet['B5'] = dict_ads['ADS']['ADDomain']
        sheet['A6'] = "Short domain name"
        sheet['B6'] = dict_ads['ADS']['Workgroup']
        sheet['A7'] = "NetBIOS machine name (override)"
        sheet['B7'] = dict_ip_information['DNS']['Hostname']
        sheet['A8'] = "Secure channel mode"
        sheet['B8'] = dict_ads['ADS']['SecureChannel']
        sheet['A9'] = "Encryption"
        sheet['B9'] = dict_ads['ADS']['Encryption']
        sheet['A10'] = "Clockskew (seconds)"
        sheet['B10'] = dict_ads['ADS']['Clockskew']
        sheet.merge_cells('A11:D11')
        sheet['A11'] = "Domain Controller"
        sheet['A11'].style = 'styleObj'
        sheet['A12'] = "Use DNS SRV lookup to obtain Domain Controller addresses"

        dc_address = {}
        kdc_address = {}
        for i in range(5):
            try:
                dc_address.update({'DC {} Address'.format(i+1): dict_ads['ADS']['DC {} Address'.format(i+1)]})
            except KeyError:
                continue
        for i in range(5):
            try:
                kdc_address.update({'KDC {} Address'.format(i+1): dict_ads['ADS']['KDC {} Address'.format(i+1)]})
                kdc_address.update({'KDC {} Port'.format(i+1): dict_ads['ADS']['KDC {} Port'.format(i+1)]})
            except KeyError:
                continue
        if not dc_address and not kdc_address:
            sheet['B12'] = "Yes"
            sheet.merge_cells('A13:D13')
            sheet['A13'] = "Kerberos Key Distribution Center"
            sheet['A13'].style = 'styleObj'
            sheet['A14'] = "Use DNS SRV lookup to obtain Kerberos Key Distribution Center addresses"
            sheet['B14'] = "Yes"
            sheet.merge_cells('A15:D15')
            sheet['A15'] = "Domain administrator credentials "
            sheet['A15'].style = 'styleObj'
            sheet['A16'] = "Username"
            sheet['B16'] = ""
            sheet['A17'] = "Password"
            sheet['B17'] = ""
        elif dc_address and not kdc_address:
            sheet['B12'] = "No"
            for i in range(5):
                sheet['A{0}'.format(i + 13)].value = "Address {}".format(i+1)
                try:
                    sheet['B{0}'.format(i+13)].value = dc_address['DC {} Address'.format(i+1)]
                except KeyError:
                    sheet['B{0}'.format(i+13)].value = ""
            sheet.merge_cells('A18:D18')
            sheet['A18'] = "Kerberos Key Distribution Center"
            sheet['A18'].style = 'styleObj'
            sheet['A19'] = "Use DNS SRV lookup to obtain Kerberos Key Distribution Center addresses"
            sheet['B19'] = "Yes"
            sheet.merge_cells('A20:D20')
            sheet['A20'] = "Domain administrator credentials "
            sheet['A20'].style = 'styleObj'
            sheet['A21'] = "Username"
            sheet['B21'] = ""
            sheet['A22'] = "Password"
            sheet['B22'] = ""
        elif not dc_address and kdc_address:
            sheet['B12'] = "Yes"
            sheet.merge_cells('A13:D13')
            sheet['A13'] = "Kerberos Key Distribution Center"
            sheet['A13'].style = 'styleObj'
            sheet['A14'] = "Use DNS SRV lookup to obtain Kerberos Key Distribution Center addresses"
            sheet['B14'] = "No"
            for i in range(5):
                sheet['A{0}'.format(i + 15)].value = "Address {}".format(i+1)
                sheet['C{0}'.format(i + 15)].value = "Port {}".format(i+1)
                try:
                    sheet['B{0}'.format(i+15)].value = kdc_address['KDC {} Address'.format(i+1)]
                except KeyError:
                    sheet['B{0}'.format(i+15)].value = ""
                try:
                    sheet['D{0}'.format(i+15)].value = kdc_address['KDC {} Port'.format(i+1)]
                except KeyError:
                    sheet['D{0}'.format(i+15)].value = ""
            sheet.merge_cells('A20:D20')
            sheet['A20'] = "Domain administrator credentials "
            sheet['A20'].style = 'styleObj'
            sheet['A21'] = "Username"
            sheet['B21'] = ""
            sheet['A22'] = "Password"
            sheet['B22'] = ""
        else:
            sheet['B12'] = "No"
            sheet['A18'] = "Use DNS SRV lookup to obtain Kerberos Key Distribution Center addresses"
            sheet['B18'] = "No"
            for i in range(5):
                sheet['A{0}'.format(i + 13)].value = "Address {}".format(i+1)
                sheet['A{0}'.format(i + 19)].value = "Address {}".format(i+1)
                sheet['C{0}'.format(i + 19)].value = "Port {}".format(i+1)
                try:
                    sheet['B{0}'.format(i+13)].value = dc_address['DC {0} Address'.format(i+1)]
                except KeyError:
                    sheet['B{0}'.format(i+13)].value = ""
                try:
                    sheet['B{0}'.format(i+19)].value = kdc_address['KDC {0} Address'.format(i+1)]
                except KeyError:
                    sheet['B{0}'.format(i+19)].value = ""
                try:
                    sheet['D{0}'.format(i+19)].value = kdc_address['KDC {0} Port'.format(i+1)]
                except KeyError:
                    sheet['D{0}'.format(i+19)].value = ""

            sheet.merge_cells('A24:D24')
            sheet['A24'] = "Domain administrator credentials "
            sheet['A24'].style = 'styleObj'
            sheet['A25'] = "Username"
            sheet['B25'] = ""
            sheet['A26'] = "Password"
            sheet['B26'] = ""

    # d) H.350 authentication from xconfiguration file
    wb.create_sheet(title="Config > Auth > Devices > H.350")
    sheet = wb.get_sheet_by_name("Config > Auth > Devices > H.350")
    sheet.column_dimensions['A'].width = 60
    sheet.column_dimensions['B'].width = 25
    sheet.merge_cells('A1:D1')


    try:
        sheet['A1'] = "H.350 directory service configuration"
        sheet['A1'].style = 'styleObj'
        sheet['A2'] = "H.350 device authentication"
        sheet['B2'] = dict_h350['H350']['Mode']
        sheet['A3'] = "Source of aliases for registration"
        sheet['B3'] = dict_h350['LDAP']['AliasOrigin']
        sheet.merge_cells('A4:D4')
        sheet['A4'] = "LDAP server configuration"
        sheet['A4'].style = 'styleObj'
        sheet['A5'] = "Server address"
        sheet['B5'] = dict_h350['H350']['LdapServerAddress']
        sheet['A6'] = "FQDN address resolution"
        sheet['B6'] = dict_h350['H350']['LdapServerAddressResolution']
        sheet['A7'] = "Port"
        if dict_h350['H350']['LdapServerAddressResolution']=="AddressRecord":
            sheet['B7'] = dict_h350['H350']['LdapServerPort']
        else:
            sheet['B7'] = ""
        sheet['A8'] = "Encryption"
        sheet['B8'] = dict_h350['H350']['LdapEncryption']
        sheet.merge_cells('A9:D9')
        sheet['A9'] = "Authentication configuration"
        sheet['A9'].style = 'styleObj'
        sheet['A10'] = "Bind DN"
        sheet['B10'] = dict_h350['H350']['BindUserDn']
        sheet['A11'] = "Bind password"
        sheet['B11'] = dict_h350['H350']['BindPassword']
        sheet.merge_cells('A12:D12')
        sheet['A12'] = "Directory configuration"
        sheet['A12'].style = 'styleObj'
        sheet['A13'] = "Base DN for devices"
        sheet['B13'] = dict_h350['H350']['DirectoryBaseDn']

    except KeyError:
        pass


    # 12) Configuration > Call Routing
    ##########################################################################
    ##########################################################################
    dict_call = xconf_to_dict(xconfiguration_file,regex_call)
    wb.create_sheet(title="Config > Call Routing")
    sheet = wb.get_sheet_by_name("Config > Call Routing")
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 25

    sheet.merge_cells('A1:B1')
    sheet['A1'] = "Configuration"
    sheet['A1'].style = 'styleObj'
    sheet['A2'] = "Call signaling optimization"
    call_optimization = dict_call['Routed']['Mode']
    if call_optimization == "Optimal":
        call_optimization = "On"
    else:
        call_optimization = "Off"
    sheet['B2'] = call_optimization
    sheet['A3'] = "Call loop detection mode"
    sheet['B3'] = dict_call['Loop']['Detection Mode']

    # 13) Configuration > Local Zone > Default Subzone
    ##########################################################################
    ##########################################################################
    dict_def_subzone = xconf_to_dict(xconfiguration_file,regex_default_subzone)
    wb.create_sheet(title="Config > LZ > Default Sub")
    sheet = wb.get_sheet_by_name("Config > LZ > Default Sub")
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 25

    sheet.merge_cells('A1:B1')
    sheet['A1'] = "Policy"
    sheet['A1'].style = 'styleObj'
    sheet['A2'] = "Registration policy"
    sheet['B2'] = dict_def_subzone['DefaultSubZone']['Registrations']
    sheet['A3'] = "Authentication policy"
    sheet['B3'] = dict_def_subzone['DefaultSubZone']['Authentication Mode']
    sheet.merge_cells('A4:B4')
    sheet['A4'] = "SIP"
    sheet['A4'].style = 'styleObj'
    sheet['A5'] = "Media encryption mode"
    sheet['B5'] = dict_def_subzone['DefaultSubZone']['SIP Media Encryption Mode']
    sheet['A6'] = "ICE support"
    sheet['B6'] = dict_def_subzone.get('DefaultSubZone').get('SIP Media ICE Support',"N/A this version of VCS")
    sheet.merge_cells('A7:B7')
    sheet['A7'] = "Total bandwidth available"
    sheet['A7'].style = 'styleObj'
    sheet['A8'] = "Bandwidth restriction"
    sheet['B8'] = dict_def_subzone['DefaultSubZone']['Bandwidth Total Mode']
    sheet['A9'] = "Total bandwidth (kbps)"
    try:
        sheet['B9'] = dict_def_subzone['DefaultSubZone']['Bandwidth Total Limit']
    except KeyError:
        sheet['B9'] = ""
    sheet.merge_cells('A10:B10')
    sheet['A10'] = "Calls into or out of the Default Subzone"
    sheet['A10'].style = 'styleObj'
    sheet['A11'] = "Bandwidth restriction"
    sheet['B11'] = dict_def_subzone['DefaultSubZone']['Bandwidth PerCall Inter Mode']
    sheet['A12'] = "Per call bandwidth limit (kbps)"
    try:
        sheet['B12'] = dict_def_subzone['DefaultSubZone']['Bandwidth PerCall Inter Limit']
    except:
        sheet['B9'] = ""
    sheet.merge_cells('A13:B13')
    sheet['A13'] = "Calls entirely within the Default Subzone"
    sheet['A13'].style = 'styleObj'
    sheet['A14'] = "Bandwidth restriction"
    sheet['B14'] = dict_def_subzone['DefaultSubZone']['Bandwidth PerCall Intra Mode']
    sheet['A15'] = "Per call bandwidth limit (kbps)"
    try:
        sheet['B15'] = dict_def_subzone['DefaultSubZone']['Bandwidth PerCall Intra Limit']
    except:
        sheet['B15'] = ""

    # 14) Configuration > Traversal Subzone
    ##########################################################################
    ##########################################################################
    wb.create_sheet(title="Config > Traversal Sub")
    sheet = wb.get_sheet_by_name("Config > Traversal Sub")
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 25

    dict_traversal_sub = xconf_to_dict(xconfiguration_file,regex_traversal_subzone)

    sheet.merge_cells('A1:B1')
    sheet['A1'] = "Ports"
    sheet['A1'].style = 'styleObj'
    sheet['A2'] = "Traversal media port start"
    sheet['A3'] = "Traversal media port end"
    sheet['B2'] = dict_traversal_media_range['Traversal Media']['Start']
    sheet['B3'] = dict_traversal_media_range['Traversal Media']['End']
    sheet.merge_cells('A4:B4')
    sheet['A4'] = "Total bandwidth available"
    sheet['A4'].style = 'styleObj'
    sheet['A5'] = "Bandwidth restriction"
    sheet['B5'] = dict_traversal_sub['TraversalSubZone']['Bandwidth Total Mode']
    sheet['A6'] = "Total bandwidth limit (kbps)"
    try:
        sheet['B6'] = dict_traversal_sub['TraversalSubZone']['Bandwidth Total Limit']
    except:
        sheet['B6'] = ""
    sheet.merge_cells('A7:B7')
    sheet['A7'] = "Calls handled by the Traversal Subzone"
    sheet['A7'].style = 'styleObj'
    sheet['A8'] = "Bandwidth restriction"
    sheet['B8'] = dict_traversal_sub['TraversalSubZone']['Bandwidth PerCall Mode']
    sheet['A9'] = "Per call bandwidth limit (kbps)"

    try:
        sheet['B9'] = dict_traversal_sub['TraversalSubZone']['Bandwidth PerCall Limit']
    except:
        sheet['B9'] = ""

    # 15) Configuration > LocalZone > Subzones
    ##########################################################################
    ##########################################################################
    wb.create_sheet(title="Config > LZ > Subzones")
    sheet = wb.get_sheet_by_name("Config > LZ > Subzones")
    sheet.column_dimensions['A'].width = 50
    dict_subzone = xconf_to_dict(xconfiguration_file,regex_subzones)

    if dict_subzone:
        sheet['A1'] = "Name"
        sheet.column_dimensions['A'].width = 25
        sheet['A1'].style = 'styleObj'
        sheet['B1'] = "Registration policy"
        sheet['B1'].style = 'styleObj'
        sheet.column_dimensions['B'].width = 25
        sheet['C1'] = "Authentication policy"
        sheet['C1'].style = 'styleObj'
        sheet.column_dimensions['C'].width = 25
        sheet['D1'] = "Media encryption mode"
        sheet['D1'].style = 'styleObj'
        sheet.column_dimensions['D'].width = 25
        sheet['E1'].style = 'styleObj'
        sheet['E1'] = "ICE support"
        sheet.column_dimensions['E'].width = 25
        sheet['F1'] = "Total Bandwidth restriction"
        sheet['F1'].style = 'styleObj'
        sheet.column_dimensions['F'].width = 25
        sheet['G1'].style = 'styleObj'
        sheet['G1'] = "Total bandwidth limit (kbps)"
        sheet.column_dimensions['G'].width = 25
        sheet['H1'].style = 'styleObj'
        sheet['H1'] = "Inter Sub Bandwidth restriction"
        sheet.column_dimensions['H'].width = 25
        sheet['I1'] = "Per call bandwidth limit (kbps)"
        sheet['I1'].style = 'styleObj'
        sheet.column_dimensions['I'].width = 25
        sheet['J1'] = "Intra Sub Bandwidth restriction"
        sheet['J1'].style = 'styleObj'
        sheet.column_dimensions['J'].width = 25
        sheet['K1'] = "Per call bandwidth limit (kbps)"
        sheet['K1'].style = 'styleObj'
        sheet.column_dimensions['K'].width = 25


        for enum,keys in enumerate(dict_subzone,2):
            sheet.cell(row = enum, column = 1).value = dict_subzone[keys]['Name']
            sheet.cell(row = enum, column = 2).value = dict_subzone[keys]['Registrations']
            sheet.cell(row = enum, column = 3).value = dict_subzone[keys]['Authentication Mode']
            sheet.cell(row = enum, column = 4).value = dict_subzone[keys]['SIP Media Encryption Mode']
            sheet.cell(row = enum, column = 5).value = dict_subzone.get(keys).get('SIP Media ICE Support',"N/A this VCS")
            sheet.cell(row = enum, column = 6).value = dict_subzone[keys]['Bandwidth Total Mode']
            sheet.cell(row = enum, column = 7).value = dict_subzone.get(keys).get('Bandwidth Total Limit',"")
            sheet.cell(row = enum, column = 8).value = dict_subzone[keys]['Bandwidth PerCall Intra Mode']
            sheet.cell(row = enum, column = 9).value = dict_subzone.get(keys).get('Bandwidth PerCall Intra Limit',"")
            sheet.cell(row = enum, column = 10).value = dict_subzone[keys]['Bandwidth PerCall Inter Mode']
            sheet.cell(row = enum, column = 11).value = dict_subzone.get(keys).get('Bandwidth PerCall Inter Limit',"")

    # 16) Configuration > LocalZone > Subzones Membership rules
    ##########################################################################
    ##########################################################################
    wb.create_sheet(title="Config > LZ > SubZ Membership")
    sheet = wb.get_sheet_by_name("Config > LZ > SubZ Membership")
    dict_subzone_membership = xconf_to_dict(xconfiguration_file,regex_membership)

    if dict_subzone_membership:
        sheet['A1'] = "Name"
        sheet['A1'].style = 'styleObj'
        sheet.column_dimensions['A'].width = 25
        sheet['B1'] = "Description"
        sheet['B1'].style = 'styleObj'
        sheet.column_dimensions['B'].width = 25
        sheet['C1'] = "Priority"
        sheet['C1'].style = 'styleObj'
        sheet.column_dimensions['C'].width = 25
        sheet['D1'] = "Type"
        sheet['D1'].style = 'styleObj'
        sheet.column_dimensions['D'].width = 25
        sheet['E1'] = "Pattern Type / Subnet address"
        sheet['E1'].style = 'styleObj'
        sheet.column_dimensions['E'].width = 25
        sheet['F1'] = "String / Prefix length"
        sheet['F1'].style = 'styleObj'
        sheet.column_dimensions['F'].width = 25
        sheet['G1'].style = 'styleObj'
        sheet.column_dimensions['G'].width = 25
        sheet['G1'] = "Target Subzone"
        sheet['H1'] = "State"
        sheet['H1'].style = 'styleObj'
        sheet.column_dimensions['H'].width = 25

        for enum, keys in enumerate(dict_subzone_membership,2):
            sheet.cell(row = enum, column = 1).value = dict_subzone_membership[keys]['Name']
            sheet.cell(row = enum, column = 2).value = dict_subzone_membership[keys]['Description']
            sheet.cell(row = enum, column = 3).value = dict_subzone_membership[keys]['Priority']
            if dict_subzone_membership[keys]['Type'] == "AliasPatternMatch":
                sheet.cell(row = enum, column = 4).value = dict_subzone_membership[keys]['Type']
                sheet.cell(row = enum, column = 5).value = dict_subzone_membership[keys]['Pattern Type']
                sheet.cell(row = enum, column = 6).value = dict_subzone_membership[keys]['Pattern String']
            elif dict_subzone_membership[keys]['Type'] == "Subnet":
                sheet.cell(row = enum, column = 4).value = dict_subzone_membership[keys]['Type']
                sheet.cell(row = enum, column = 5).value = dict_subzone_membership[keys]['Subnet Address']
                sheet.cell(row = enum, column = 6).value = dict_subzone_membership[keys]['Subnet PrefixLength']
            sheet.cell(row = enum, column = 7).value = dict_subzone_membership[keys]['SubZoneName']
            sheet.cell(row = enum, column = 8).value = dict_subzone_membership[keys]['State']


    # 17) Configuration > Zones > Zones > Default Zone
    ##########################################################################
    ##########################################################################
    wb.create_sheet(title="Config > DefaultZone")
    sheet = wb.get_sheet_by_name("Config > DefaultZone")
    dict_def_zone = xconf_to_dict(xconfiguration_file,regex_def_zone)

    sheet.column_dimensions['A'].width = 30
    sheet.merge_cells('A1:B1')
    sheet['A1'] = "Policy"
    sheet['A1'].style = 'styleObj'
    sheet['A2'] = "Authentication policy"
    sheet['B2'] = dict_def_zone['DefaultZone']['Authentication Mode']
    sheet.merge_cells('A3:B3')
    sheet['A3'] = "SIP"
    sheet['A3'].style = 'styleObj'
    sheet['A4'] = "Media encryption mode"
    sheet['B4'] = dict_def_zone['DefaultZone']['SIP Media Encryption Mode']
    sheet['A5'] = "ICE support"
    sheet['B5'] = dict_def_zone.get('DefaultZone').get('SIP Media ICE Support',"N/A to this VCS")
    sheet['A6'] = "Use Default Zone access rules on port"
    tls_verify_mode = dict_def_zone['DefaultZone']['SIP TLS Verify Mode']
    if tls_verify_mode == "On":
        tls_verify_mode = "TLS (5061) and MTLS (5062)"
    else:
        tls_verify_mode = "MTLS Only (5062)"

    sheet['B6'] = tls_verify_mode


    # 18) Configuration > Zones > Zones
    ##########################################################################
    ##########################################################################
    dict_zone_type = xconf_to_dict(xconfiguration_file,regex_zone_type)
    dict_zone_config = xconf_to_dict(xconfiguration_file,regex_zone_config)

    for keys in sorted(dict_zone_type):
        zone_title = dict_zone_config[keys]['Name'][:31]
        wb.create_sheet(title = zone_title)
        sheet = wb.get_sheet_by_name(zone_title)
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 30
        sheet.merge_cells('A1:B1')
        sheet['A1'] = "Configuration"
        sheet['A1'].style = 'styleObj'
        sheet['A2'] = "Name"
        sheet['B2'] = dict_zone_config[keys]['Name']
        sheet['A3'] = "Type"
        sheet['B3'] = dict_zone_type[keys]['Type']
        sheet['A4'] = "Hop count"
        sheet['B4'] = dict_zone_config[keys]['HopCount']

        if dict_zone_type[keys]['Type'] == "TraversalClient":
            sheet.merge_cells('A5:B5')
            sheet['A5'] = "Connection credentials"
            sheet['A5'].style = 'styleObj'
            sheet['A6'] = "Username"
            sheet['A7'] = "Password"
            sheet['B6'] = dict_zone_config[keys]['TraversalClient Authentication UserName']
            sheet['B7'] = dict_zone_config[keys]['TraversalClient Authentication Password']
            try:
                if dict_zone_config[keys]['TraversalClient Collaboration Edge'] == "Off": # Looking to whether the zone is Traversal or UC Traversal
                    sheet.merge_cells('A8:F8')
                    sheet['A8'].style = 'styleObj'
                    sheet['A8'] = "H.323"
                    sheet['A9'] = "Mode"
                    sheet['B9'] = dict_zone_config[keys]['H323 Mode']
                    sheet['A10'] = "Protocol"
                    sheet['B10'] = dict_zone_config[keys]['TraversalClient H323 Protocol']
                    sheet['A11'] = "Port"
                    sheet['B11'] = dict_zone_config[keys]['TraversalClient H323 Port']
                    sheet.merge_cells('A12:F12')
                    sheet['A12'].style = 'styleObj'
                    sheet['A12'] = "SIP"
                    sheet['A13'] = "Mode"
                    sheet['B13'] = dict_zone_config[keys]['SIP Mode']
                    sheet['A14'] = "Port"
                    sheet['B14'] = dict_zone_config[keys]['TraversalClient SIP Port']
                    sheet['A15'] = "Transport"
                    sheet['B15'] = dict_zone_config[keys]['TraversalClient SIP Transport']
                    sheet['A16'] = "TLS verify mode"
                    sheet['B16'] = dict_zone_config[keys]['TraversalClient SIP TLS Verify Mode']
                    sheet['A17'] = "Accept proxied registrations"
                    sheet['B17'] = dict_zone_config[keys]['TraversalClient Registrations']
                    sheet['A18'] = "Media encryption mode"
                    sheet['B18'] = dict_zone_config[keys]['TraversalClient SIP Media Encryption Mode']
                    sheet['A19'] = "ICE support"
                    sheet['B19'] = dict_zone_config[keys]['TraversalClient SIP Media ICE Support']
                    sheet['A20'] = "SIP Poison mode"
                    sheet['B20'] = dict_zone_config[keys]['TraversalClient SIP Poison Mode']
                    sheet['A21'] = "Preloaded SIP routes support"
                    try:
                        sheet['B21'] = dict_zone_config[keys]['TraversalClient SIP PreloadedSipRoutes Accept']
                    except KeyError:
                        sheet['B21'] = "Not defined on this version of VCS"
                    sheet['A22'] = "SIP parameter preservation"
                    try:
                        sheet['B22'] = dict_zone_config[keys]['TraversalClient SIP ParameterPreservation Mode']
                    except KeyError:
                        sheet['B22'] = "Not defined on this version of VCS"
                    sheet.merge_cells('A23:F23')
                    sheet['A23'].style = 'styleObj'
                    sheet['A23'] = "Authentication"
                    sheet['A24'] = "Authentication policy"
                    sheet['B24'] = dict_zone_config[keys]['TraversalClient Authentication Mode']
                    sheet['A25'] = "Accept delegated credential checks"
                    sheet['B25'] = dict_zone_config[keys]['TraversalClient Accept Delegated Credential Checks']
                    sheet.merge_cells('A26:F26')
                    sheet['A26'].style = 'styleObj'
                    sheet['A26'] = "Client settings"
                    sheet['A27'] = "Retry interval"
                    sheet['B27'] = dict_zone_config[keys]['TraversalClient RetryInterval']
                    sheet.merge_cells('A28:F28')
                    sheet['A28'].style = 'styleObj'
                    sheet['A28'] = "Location"
                    for peers in range(6):
                        sheet['A{0}'.format(29 + peers)].value = "Peer {} address".format(peers + 1)
                        sheet['B{0}'.format(29 + peers)].value = dict_zone_config[keys]['TraversalClient Peer {} Address'.format(peers + 1)]

                elif dict_zone_config[keys]['TraversalClient Collaboration Edge'] == "On":
                    sheet.merge_cells('A8:F8')
                    sheet['A8'].style = 'styleObj'
                    sheet['A8'] = "SIP"
                    sheet['A9'] = "Port"
                    sheet['B9'] = dict_zone_config[keys]['TraversalClient SIP Port']
                    sheet['A10'] = "Accept proxied registrations"
                    sheet['B10'] = dict_zone_config[keys]['TraversalClient Registrations']
                    sheet['A11'] = "ICE support"
                    sheet['B11'] = dict_zone_config[keys]['TraversalClient SIP Media ICE Support']
                    sheet['A12'] = "SIP Poison mode"
                    sheet['B12'] = dict_zone_config[keys]['TraversalClient SIP Poison Mode']
                    sheet['A13'] = "Preloaded SIP routes support"
                    sheet['B13'] = dict_zone_config[keys]['TraversalClient SIP PreloadedSipRoutes Accept']
                    sheet['A14'] = "SIP parameter preservation"
                    sheet['B14'] = dict_zone_config[keys]['TraversalClient SIP ParameterPreservation Mode']
                    sheet.merge_cells('A15:F15')
                    sheet['A15'].style = 'styleObj'
                    sheet['A15'] = "Authentication"
                    sheet['A16'] = "Authentication policy"
                    sheet['B16'] = dict_zone_config[keys]['TraversalClient Authentication Mode']
                    sheet['A17'] = "Accept delegated credential checks"
                    sheet['B17'] = dict_zone_config[keys]['TraversalClient Accept Delegated Credential Checks']
                    sheet.merge_cells('A18:F18')
                    sheet['A18'].style = 'styleObj'
                    sheet['A18'] = "Client settings"
                    sheet['A19'] = "Retry interval"
                    sheet['B19'] = dict_zone_config[keys]['TraversalClient RetryInterval']
                    sheet.merge_cells('A20:F20')
                    sheet['A20'].style = 'styleObj'
                    sheet['A20'] = "Location"
                    for peers in range(6):
                        sheet['A{0}'.format(21 + peers)].value = "Peer {} address".format(peers + 1)
                        sheet['B{0}'.format(21 + peers)].value = dict_zone_config[keys]['TraversalClient Peer {} Address'.format(peers + 1)]
            except:
                sheet.merge_cells('A8:F8')
                sheet['A8'].style = 'styleObj'
                sheet['A8'] = "H.323"
                sheet['A9'] = "Mode"
                sheet['B9'] = dict_zone_config[keys]['H323 Mode']
                sheet['A10'] = "Protocol"
                sheet['B10'] = dict_zone_config[keys]['TraversalClient H323 Protocol']
                sheet['A11'] = "Port"
                sheet['B11'] = dict_zone_config[keys]['TraversalClient H323 Port']
                sheet.merge_cells('A12:F12')
                sheet['A12'].style = 'styleObj'
                sheet['A12'] = "SIP"
                sheet['A13'] = "Mode"
                sheet['B13'] = dict_zone_config[keys]['SIP Mode']
                sheet['A14'] = "Port"
                sheet['B14'] = dict_zone_config[keys]['TraversalClient SIP Port']
                sheet['A15'] = "Transport"
                sheet['B15'] = dict_zone_config[keys]['TraversalClient SIP Transport']
                sheet['A16'] = "TLS verify mode"
                sheet['B16'] = dict_zone_config[keys]['TraversalClient SIP TLS Verify Mode']
                sheet['A17'] = "Accept proxied registrations"
                sheet['B17'] = dict_zone_config[keys]['TraversalClient Registrations']
                sheet['A18'] = "Media encryption mode"
                sheet['B18'] = dict_zone_config[keys]['TraversalClient SIP Media Encryption Mode']
                sheet['A19'] = "ICE support"
                sheet['B19'] = dict_zone_config.get(keys).get('TraversalClient SIP Media ICE Support',"N/A")
                sheet['A20'] = "SIP Poison mode"
                sheet['B20'] = dict_zone_config[keys]['TraversalClient SIP Poison Mode']
                sheet['A21'] = "Preloaded SIP routes support"
                try:
                    sheet['B21'] = dict_zone_config[keys]['TraversalClient SIP PreloadedSipRoutes Accept']
                except KeyError:
                    sheet['B21'] = "Not defined on this version of VCS"
                sheet['A22'] = "SIP parameter preservation"
                try:
                    sheet['B22'] = dict_zone_config[keys]['TraversalClient SIP ParameterPreservation Mode']
                except KeyError:
                    sheet['B22'] = "Not defined on this version of VCS"
                sheet.merge_cells('A23:F23')
                sheet['A23'].style = 'styleObj'
                sheet['A23'] = "Authentication"
                sheet['A24'] = "Authentication policy"
                sheet['B24'] = dict_zone_config[keys]['TraversalClient Authentication Mode']
                sheet['A25'] = "Accept delegated credential checks"
                sheet['B25'] = dict_zone_config.get(keys).get('TraversalClient Accept Delegated Credential Checks',"N/A to this VCS")
                sheet.merge_cells('A26:F26')
                sheet['A26'].style = 'styleObj'
                sheet['A26'] = "Client settings"
                sheet['A27'] = "Retry interval"
                sheet['B27'] = dict_zone_config[keys]['TraversalClient RetryInterval']
                sheet.merge_cells('A28:F28')
                sheet['A28'].style = 'styleObj'
                sheet['A28'] = "Location"
                for peers in range(6):
                    sheet['A{0}'.format(29 + peers)].value = "Peer {} address".format(peers + 1)
                    sheet['B{0}'.format(29 + peers)].value = dict_zone_config[keys][
                        'TraversalClient Peer {} Address'.format(peers + 1)]

        elif dict_zone_type[keys]['Type'] == "Neighbor":
            sheet.merge_cells('A5:F5')
            sheet['A5'].style = 'styleObj'
            sheet['A5'] = "H.323"
            sheet['A6'] = "Mode"
            sheet['B6'] = dict_zone_config[keys]['H323 Mode']
            sheet['A7'] = "Port"
            sheet['B7'] = dict_zone_config[keys]['Neighbor H323 Port']
            sheet.merge_cells('A8:F8')
            sheet['A8'].style = 'styleObj'
            sheet['A8'] = "SIP"
            sheet['A9'] = "Mode"
            sheet['B9'] = dict_zone_config[keys]['SIP Mode']
            sheet['A10'] = "Port"
            sheet['B10'] = dict_zone_config[keys]['Neighbor SIP Port']
            sheet['A11'] = "Transport"
            sheet['B11'] = dict_zone_config[keys]['Neighbor SIP Transport']
            sheet['A12'] = "TLS verify mode"
            sheet['B12'] = dict_zone_config[keys]['Neighbor SIP TLS Verify Mode']
            sheet['A13'] = "Accept proxied registrations"
            sheet['B13'] = dict_zone_config[keys]['Neighbor SIP TLS Verify Mode']
            sheet['A14'] = "Media encryption mode"
            sheet['B14'] = dict_zone_config[keys]['Neighbor SIP Media Encryption Mode']
            sheet['A15'] = "ICE support"
            sheet['B15'] = dict_zone_config.get(keys).get('Neighbor SIP Media ICE Support',"N/A")
            sheet['A16'] = "Preloaded SIP routes support"
            sheet['B16'] = dict_zone_config.get(keys).get("Neighbor SIP PreloadedSipRoutes Accept","Off")

            sheet.merge_cells('A17:F17')
            sheet['A17'].style = 'styleObj'
            sheet['A17'] = "Authentication"
            sheet['A18'] = "Authentication policy"
            sheet['B18'] = dict_zone_config[keys]['Neighbor Authentication Mode']
            sheet['A19'] = "SIP authentication trust mode"
            sheet['B19'] = dict_zone_config[keys]['Neighbor SIP Authentication Trust Mode']
            sheet.merge_cells('A20:F20')
            sheet['A20'].style = 'styleObj'
            sheet['A20'] = "Location"
            for peers in range(6):
                sheet['A{0}'.format(21 + peers)].value = "Peer {} address".format(peers + 1)
                sheet['B{0}'.format(21 + peers)].value = dict_zone_config[keys]['Neighbor Peer {} Address'.format(peers + 1)]

            sheet.merge_cells('A27:F27')
            sheet['A27'].style = 'styleObj'
            sheet['A27'] = "Advanced"
            sheet['A28'] = "Zone profile"
            sheet['B28'] = dict_zone_config[keys]['Neighbor ZoneProfile']
            sheet['A29'] = "Monitor peer status"
            sheet['B29'] = dict_zone_config.get(keys).get('Neighbor Monitor',"N/A")
            sheet['A30'] = "Call signaling routed mode"
            sheet['B30'] = dict_zone_config.get(keys).get('Neighbor SignalingRouting Mode',"N/A")
            sheet['A31'] = "Automatically respond to H.323 searches"
            sheet['B31'] = dict_zone_config.get(keys).get('Neighbor H323 SearchAutoResponse',"N/A")
            sheet['A32'] = "Automatically respond to SIP searches"
            sheet['B32'] = dict_zone_config.get(keys).get('Neighbor Interworking SIP Search Strategy',"N/A")
            sheet['A33'] = "Send empty INVITE for interworked calls"
            sheet['B33'] = dict_zone_config.get(keys).get('Neighbor Interworking SIP EmptyInviteAllowed',"N/A")
            sheet['A34'] = "SIP parameter preservation"
            sheet['B34'] = dict_zone_config.get(keys).get("Neighbor SIP ParameterPreservation Mode","N/A")
            sheet['A35'] = "SIP poison mode"
            sheet['B35'] = dict_zone_config.get(keys).get('Neighbor SIP Poison Mode',"Off")
            sheet['A36'] = "SIP encryption mode"
            sheet['B36'] = dict_zone_config.get(keys).get('Neighbor SIP Media Encryption Mode',"N/A")
            sheet['A37'] = "SIP REFER mode"
            sheet['B37'] = dict_zone_config.get(keys).get('Neighbor SIP B2BUA Refer Mode',"N/A")
            sheet['A38'] = "SIP multiparty MIME strip mode"
            sheet['B38'] = dict_zone_config.get(keys).get('Neighbor SIP MIME Strip Mode',"N/A")
            sheet['A39'] = "SIP UPDATE strip mode"
            sheet['B39'] = dict_zone_config.get(keys).get('Neighbor SIP UPDATE Strip Mode',"N/A")
            sheet['A40'] = "Interworking SIP search strategy"
            sheet['B40'] = dict_zone_config.get(keys).get('Neighbor Interworking SIP Search Strategy',"N/A")
            sheet['A41'] = "SIP UDP/BFCP filter mode"
            sheet['B41'] = dict_zone_config.get(keys).get('Neighbor SIP UDP BFCP Filter Mode',"N/A")
            sheet['A42'] = "SIP UDP/IX filter mode"
            sheet['B42'] = dict_zone_config.get(keys).get('Neighbor SIP UDP IX Filter Mode',"N/A")
            sheet['A43'] = "SIP record route address type"
            sheet['B43'] = dict_zone_config.get(keys).get('Neighbor SIP Record Route Address Type',"N/A")
            sheet['A44'] = "SIP Proxy-Require header strip list"
            sheet['B44'] = dict_zone_config.get(keys).get('Neighbor SIP ProxyRequire Strip List',"N/A")

    ### DNS Zone
        elif dict_zone_type[keys]['Type'] == "DNS":
            sheet.merge_cells('A5:F5')
            sheet['A5'].style = 'styleObj'
            sheet['A5'] = "H.323"
            sheet['A6'] = "Mode"
            sheet['B6'] = dict_zone_config[keys]['H323 Mode']
            sheet.merge_cells('A7:F7')
            sheet['A7'].style = 'styleObj'
            sheet['A7'] = "SIP"
            sheet['A8'] = "Mode"
            sheet['B8'] = dict_zone_config[keys]['SIP Mode']
            sheet['A9'] = "TLS verify mode"
            sheet['B9'] = dict_zone_config[keys]['DNS SIP TLS Verify Mode']
            sheet['A10'] = "TLS verify subject name"
            sheet['B10'] = dict_zone_config[keys]['DNS SIP TLS Verify Subject Name']
            sheet['A11'] = "TLS verify inbound mapping"
            try:
                sheet['B11'] = dict_zone_config[keys]['DNS SIP TLS Verify InboundClassification']
            except KeyError:
                sheet['B11'] = "N/A for this version of VCS"
            sheet['A12'] = "Fallback transport protocol"
            sheet['B12'] = dict_zone_config[keys]['DNS SIP Default Transport']
            sheet['A13'] = "Media encryption mode"
            sheet['B13'] = dict_zone_config[keys]['DNS SIP Media Encryption Mode']
            sheet['A14'] = "ICE support"
            sheet['B14'] = dict_zone_config[keys]['DNS SIP Media ICE Support']
            sheet['A15'] = "Preloaded SIP routes support"
            try:
                sheet['B15'] = dict_zone_config[keys]['DNS SIP PreloadedSipRoutes Accept']
            except KeyError:
                sheet['B15'] = "N/A for this version of VCS"
            sheet['A16'] = "Modify DNS request"
            try:
                sheet['B16'] = dict_zone_config[keys]['DNS SIP DnsOverride Override']
            except KeyError:
                sheet['B16'] = "N/A for this version of VCS"
            sheet['A17'] = "Domain to search for"
            try:
                sheet['B17'] = dict_zone_config[keys]['DNS SIP DnsOverride Name']
            except KeyError:
                sheet['B17'] = "N/A for this version of VCS"
            sheet.merge_cells('A18:F18')
            sheet['A18'].style = 'styleObj'
            sheet['A18'] = "Authentication"
            sheet['A19'] = "SIP authentication trust mode"
            try:
                sheet['B19'] = dict_zone_config[keys]['DNS SIP Authentication Trust Mode']
            except KeyError:
                sheet['B19'] = "N/A for this version of VCS"
            sheet.merge_cells('A20:F20')
            sheet['A20'].style = 'styleObj'
            sheet['A20'] = "Advanced"
            sheet['A21'] = "Include address record"
            sheet['B21'] = dict_zone_config[keys]['DNS IncludeAddressRecord']
            sheet['A22'] = "Zone profile"
            sheet['B22'] = dict_zone_config[keys]['DNS ZoneProfile']
            sheet['A23'] = "Automatically respond to SIP searches"
            sheet['B23'] = dict_zone_config[keys]['DNS SIP SearchAutoResponse']
            sheet['A24'] = "Send empty INVITE for interworked calls"
            sheet['B24'] = dict_zone_config[keys]['DNS Interworking SIP EmptyInviteAllowed']
            sheet['A25'] = "SIP parameter preservation"
            try:
                sheet['B25'] = dict_zone_config[keys]['DNS SIP ParameterPreservation Mode']
            except KeyError:
                sheet['B25'] = "N/A for this version of VCS"
            sheet['A26'] = "SIP poison mode"
            sheet['B26'] = dict_zone_config[keys]['DNS SIP Poison Mode']
            sheet['A27'] = "SIP UDP/BFCP filter mode"
            sheet['B27'] = dict_zone_config[keys]['DNS SIP UDP BFCP Filter Mode']
            sheet['A28'] = "SIP UDP/IX filter mode"
            sheet['B28'] = dict_zone_config[keys]['DNS SIP UDP IX Filter Mode']
            sheet['A29'] = "SIP record route address type"
            sheet['B29'] = dict_zone_config[keys]['DNS SIP Record Route Address Type']
    ### ENUM Zone
        elif dict_zone_type[keys]['Type'] == "ENUM":
            sheet.merge_cells('A5:F5')
            sheet['A5'].style = 'styleObj'
            sheet['A5'] = "DNS settings"
            sheet['A6'] = "DNS suffix"
            sheet['B6'] = dict_zone_config[keys]['ENUM DNSSuffix']
            sheet.merge_cells('A7:F7')
            sheet['A7'].style = 'styleObj'
            sheet['A7'] = "H.323"
            sheet['A8'] = "Mode"
            sheet['B8'] = dict_zone_config[keys]['H323 Mode']
            sheet.merge_cells('A9:F9')
            sheet['A9'].style = 'styleObj'
            sheet['A9'] = "SIP"
            sheet['A10'] = "Mode"
            sheet['B10'] = dict_zone_config[keys]['SIP Mode']

    ### Traversal Server Zone
        elif dict_zone_type[keys]['Type'] == "TraversalServer":
            sheet.merge_cells('A5:B5')
            sheet['A5'] = "Connection credentials"
            sheet['A5'].style = 'styleObj'
            sheet['A6'] = "Username"
            sheet['A7'] = "Password"
            sheet['B6'] = dict_zone_config[keys]['TraversalServer Authentication UserName']
            sheet['B7'] = ""
            if dict_zone_config[keys]['TraversalServer Collaboration Edge'] == "Off": # Looking to whether the zone is Traversal or UC Traversal
                sheet.merge_cells('A8:F8')
                sheet['A8'].style = 'styleObj'
                sheet['A8'] = "H.323"
                sheet['A9'] = "Mode"
                sheet['B9'] = dict_zone_config[keys]['H323 Mode']
                sheet['A10'] = "Protocol"
                sheet['B10'] = dict_zone_config[keys]['TraversalServer H323 Protocol']
                sheet['A11'] = "Port"
                sheet['B11'] = dict_zone_config[keys]['TraversalServer H323 Port']
                sheet['A12'] = "H.460.19 demultiplexing mode"
                sheet['B12'] = dict_zone_config[keys]['TraversalServer H323 H46019 Demultiplexing Mode']
                sheet.merge_cells('A13:F13')
                sheet['A13'].style = 'styleObj'
                sheet['A13'] = "SIP"
                sheet['A14'] = "Mode"
                sheet['B14'] = dict_zone_config[keys]['SIP Mode']
                sheet['A15'] = "Port"
                sheet['B15'] = dict_zone_config[keys]['TraversalServer SIP Port']
                sheet['A16'] = "Transport"
                sheet['B16'] = dict_zone_config[keys]['TraversalServer SIP Transport']
                sheet['A17'] = "TLS verify mode"
                sheet['B17'] = dict_zone_config[keys]['TraversalServer SIP TLS Verify Mode']
                sheet['A18'] = "TLS verify subject name"
                sheet['B18'] = dict_zone_config[keys]['TraversalServer SIP TLS Verify Subject Name']
                sheet['A19'] = "Accept proxied registrations"
                sheet['B19'] = dict_zone_config[keys]['TraversalServer Registrations']
                sheet['A20'] = "Media encryption mode"
                sheet['B20'] = dict_zone_config[keys]['TraversalServer SIP Media Encryption Mode']
                sheet['A21'] = "ICE support"
                sheet['B21'] = dict_zone_config[keys]['TraversalServer SIP Media ICE Support']
                sheet['A22'] = "SIP Poison mode"
                sheet['B22'] = dict_zone_config[keys]['TraversalServer SIP Poison Mode']
                sheet['A23'] = "Preloaded SIP routes support"
                try:
                    sheet['B23'] = dict_zone_config[keys]['TraversalServer SIP PreloadedSipRoutes Accept']
                except KeyError:
                    sheet['B23'] = "N/A for this version of VCS"
                sheet['A24'] = "SIP parameter preservation"
                try:
                    sheet['B24'] = dict_zone_config[keys]['TraversalServer SIP ParameterPreservation Mode']
                except KeyError:
                    sheet['B24'] = "N/A for this version of VCS"
                sheet.merge_cells('A25:F25')
                sheet['A25'].style = 'styleObj'
                sheet['A25'] = "Authentication"
                sheet['A26'] = "Authentication policy"
                sheet['B26'] = dict_zone_config[keys]['TraversalServer Authentication Mode']
                sheet.merge_cells('A27:F27')
                sheet['A27'].style = 'styleObj'
                sheet['A27'] = "UDP / TCP probes"
                sheet['A28'] = "UDP retry interval"
                sheet['B28'] = dict_zone_config[keys]['TraversalServer UDPProbe RetryInterval']
                sheet['A29'] = "UDP retry count"
                sheet['B29'] = dict_zone_config[keys]['TraversalServer UDPProbe RetryCount']
                sheet['A30'] = "UDP keep alive interval"
                sheet['B30'] = dict_zone_config[keys]['TraversalServer UDPProbe KeepAliveInterval']
                sheet['A31'] = "TCP retry interval"
                sheet['B31'] = dict_zone_config[keys]['TraversalServer TCPProbe RetryInterval']
                sheet['A32'] = "TCP retry count"
                sheet['B32'] = dict_zone_config[keys]['TraversalServer TCPProbe RetryCount']
                sheet['A33'] = "TCP keep alive interval"
                sheet['B33'] = dict_zone_config[keys]['TraversalServer TCPProbe KeepAliveInterval']

            elif dict_zone_config[keys]['TraversalServer Collaboration Edge'] == "On":
                sheet.merge_cells('A8:F8')
                sheet['A8'].style = 'styleObj'
                sheet['A8'] = "SIP"
                sheet['A9'] = "Mode"
                sheet['B9'] = dict_zone_config[keys]['SIP Mode']
                sheet['A10'] = "Port"
                sheet['B10'] = dict_zone_config[keys]['TraversalServer SIP Port']
                sheet['A11'] = "Transport"
                sheet['B11'] = dict_zone_config[keys]['TraversalServer SIP Transport']
                sheet['A12'] = "TLS verify mode"
                sheet['B12'] = dict_zone_config[keys]['TraversalServer SIP TLS Verify Mode']
                sheet['A13'] = "TLS verify subject name"
                sheet['B13'] = dict_zone_config[keys]['TraversalServer SIP TLS Verify Subject Name']
                sheet['A14'] = "Accept proxied registrations"
                sheet['B14'] = dict_zone_config[keys]['TraversalServer Registrations']
                sheet['A15'] = "Media encryption mode"
                sheet['B15'] = dict_zone_config[keys]['TraversalServer SIP Media Encryption Mode']
                sheet['A16'] = "ICE support"
                sheet['B16'] = dict_zone_config[keys]['TraversalServer SIP Media ICE Support']
                sheet['A17'] = "SIP Poison mode"
                sheet['B17'] = dict_zone_config[keys]['TraversalServer SIP Poison Mode']
                sheet['A18'] = "Preloaded SIP routes support"
                sheet['B18'] = dict_zone_config[keys]['TraversalServer SIP PreloadedSipRoutes Accept']
                sheet['A19'] = "SIP parameter preservation"
                sheet['B19'] = dict_zone_config[keys]['TraversalServer SIP ParameterPreservation Mode']
                sheet.merge_cells('A20:F20')
                sheet['A20'].style = 'styleObj'
                sheet['A20'] = "Authentication"
                sheet['A21'] = "Authentication policy"
                sheet['B21'] = dict_zone_config[keys]['TraversalServer Authentication Mode']
                sheet.merge_cells('A22:F22')
                sheet['A22'].style = 'styleObj'
                sheet['A22'] = "UDP / TCP probes"
                sheet['A23'] = "UDP retry interval"
                sheet['B23'] = dict_zone_config[keys]['TraversalServer UDPProbe RetryInterval']
                sheet['A24'] = "UDP retry count"
                sheet['B24'] = dict_zone_config[keys]['TraversalServer UDPProbe RetryCount']
                sheet['A25'] = "UDP keep alive interval"
                sheet['B25'] = dict_zone_config[keys]['TraversalServer UDPProbe KeepAliveInterval']
                sheet['A26'] = "TCP retry interval"
                sheet['B26'] = dict_zone_config[keys]['TraversalServer TCPProbe RetryInterval']
                sheet['A27'] = "TCP retry count"
                sheet['B27'] = dict_zone_config[keys]['TraversalServer TCPProbe RetryCount']
                sheet['A28'] = "TCP keep alive interval"
                sheet['B28'] = dict_zone_config[keys]['TraversalServer TCPProbe KeepAliveInterval']


    # 19) Configuration > Domains
    ##########################################################################
    ##########################################################################
    wb.create_sheet(title="Configuration > Domains")
    sheet = wb.get_sheet_by_name("Configuration > Domains")
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 30
    sheet.column_dimensions['E'].width = 30
    sheet.column_dimensions['F'].width = 30
    sheet.column_dimensions['G'].width = 30
    sheet.merge_cells('A1:G1')
    sheet['A1'] = "Domains"
    sheet['A1'].style = 'styleObj'
    sheet['A2'] = "Index"
    sheet['B2'] = "Domain name"
    sheet['C2'] = "VCS registrations"
    sheet['D2'] = "Unified CM registrations"
    sheet['E2'] = "IM and Presence Service"
    sheet['F2'] = "XMPP Federation"
    sheet['G2'] = "Jabber Guest"

    for enum, key in enumerate(sorted(dict_sip_domain), 3):
        sheet['A{0}'.format(enum)] = key
        sheet['B{0}'.format(enum)] = dict_sip_domain[key]['Name']
        sheet['C{0}'.format(enum)] = dict_sip_domain[key]['Sip']
        sheet['D{0}'.format(enum)] = dict_sip_domain[key]['Edgesip']
        sheet['E{0}'.format(enum)] = dict_sip_domain[key]['Edgexmpp']
        sheet['F{0}'.format(enum)] = dict_sip_domain[key]['Xmppfederation']
        sheet['G{0}'.format(enum)] = dict_sip_domain[key]['Edgejabberc']

    ##########################################################################
    ##########################################################################
    ###################        MRA information            ####################
    ##########################################################################
    ##########################################################################
    try:
        wb.create_sheet(title="Configuration > UC > Conf")
        sheet = wb.get_sheet_by_name("Configuration > UC > Conf")
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 30
        dict_collab_edge = xconf_to_dict(xconfiguration_file,regex_collab_edge)
        dict_colledge_deployments =  xconf_to_dict(xconfiguration_file,regex_collab_edge_deployments)
        sheet.merge_cells('A1:B1')
        sheet['A1'].style = 'styleObj'
        sheet['A1'] = "Configuration - Use screenshot to complete information"
        sheet['A2'] = "Unified Communications mode"
        if dict_collab_edge['CollaborationEdge']['Enabled'] == "Off":
            sheet['B2'] =  dict_collab_edge['CollaborationEdge']['Enabled']
        elif dict_collab_edge['CollaborationEdge']['Enabled'] == "On" and dict_collab_edge['CollaborationEdge']['JabbercEnabled']== "Off":
            sheet['B2'] = "Mobile and remote access"
            try:
                sheet.merge_cells('A3:B3')
                sheet['A3'].style = 'styleObj'
                sheet['A3'] = "Single Sign-On support"
                sheet['A4'] = "Single Sign-On support"
                sheet['B4'] =  dict_collab_edge['CollaborationEdge']['SsoEnabled']
            except:
                logging.debug("SSO not available for Collaboration Edge")

            # Deployment information
            wb.create_sheet(title="Conf > UC > Deployments")
            sheet = wb.get_sheet_by_name("Conf > UC > Deployments")
            sheet['A1'].style = 'styleObj'
            sheet['A1'] = "Deployments"
            sheet['A2'] = "Deployment name"
            sheet['B2'] = "Domains"
            sheet['C2'] = "IM and Presence Service nodes"
            sheet['D2'] = "Unified CM servers"
            sheet['E2'] = "Configure Unity Connection servers"
            for enum, keys in enumerate(sorted(dict_colledge_deployments),3):
                sheet['A{0}'.format(enum)].value = dict_colledge_deployments[keys]['UserReadableName']
        else:
            sheet['B2'] = "Jabber Guest services"
    except KeyError:
        pass
    # The rest of the information is not contained within xconfiguration


    # 20) Configuration > Dial Plan
    ##########################################################################
    ##########################################################################
    wb.create_sheet(title="Configuration > Dial Plan")
    sheet = wb.get_sheet_by_name("Configuration > Dial Plan")
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 30

    sheet.merge_cells('A1:B1')
    sheet['A1'].style = 'styleObj'
    sheet['A1'] = "Configuration"
    sheet['A2'] = "Calls to unknown IP addresses"
    sheet['A3'] = "Fallback alias"
    sheet['B2'] = dict_call['Services']['CallsToUnknownIPAddresses']
    sheet['B3'] = dict_call['Services']['Fallback Alias']

    # 21) Configuration > Transforms
    ##########################################################################
    ##########################################################################

    wb.create_sheet(title="Configuration > Transforms")
    sheet = wb.get_sheet_by_name("Configuration > Transforms")
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 30
    sheet.column_dimensions['E'].width = 30
    sheet.column_dimensions['F'].width = 30
    sheet.column_dimensions['G'].width = 30

    dict_transforms = xconf_to_dict(xconfiguration_file,regex_transform)

    sheet.merge_cells('A1:G1')
    sheet['A1'] = "Transforms"
    sheet['A2'] = "Priority"
    sheet['B2'] = "State"
    sheet['C2'] = "Description"
    sheet['D2'] = "Pattern type"
    sheet['E2'] = "Pattern string"
    sheet['F2'] = "Pattern behavior"
    sheet['G2'] = "Replace string"

    for enum, keys in enumerate(dict_transforms,3):
        sheet['A{0}'.format(enum)].value = dict_transforms[keys]['Priority']
        sheet['B{0}'.format(enum)].value = dict_transforms[keys]['State']
        sheet['C{0}'.format(enum)].value = dict_transforms[keys]['Description']
        sheet['D{0}'.format(enum)].value = dict_transforms[keys]['Pattern Type']
        sheet['E{0}'.format(enum)].value = dict_transforms[keys]['Pattern String']
        sheet['F{0}'.format(enum)].value = dict_transforms[keys]['Pattern Behavior']
        sheet['G{0}'.format(enum)].value = dict_transforms[keys]['Pattern Replace']

    # 22) Configuration > Search Rule
    ##########################################################################
    ##########################################################################
    wb.create_sheet(title="Configuration > Search rules")
    sheet = wb.get_sheet_by_name("Configuration > Search rules")
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 20
    sheet.column_dimensions['I'].width = 20
    sheet.column_dimensions['J'].width = 20
    sheet.column_dimensions['K'].width = 20
    sheet.column_dimensions['L'].width = 20
    sheet.column_dimensions['M'].width = 20
    sheet.column_dimensions['N'].width = 20
    sheet.column_dimensions['O'].width = 20

    dict_search_rule =  xconf_to_dict(xconfiguration_file,regex_search)

    sheet.merge_cells('A1:O1')
    sheet['A1'] = "Search rules"
    sheet['A2'] = "Priority"
    sheet['B2'] = "Rule name"
    sheet['C2'] = "Description"
    sheet['D2'] = "Protocol"
    sheet['E2'] = "Source"
    sheet['F2'] = "Source name"
    sheet['G2'] = "Authentication required"
    sheet['H2'] = "Mode"
    sheet['I2'] = "Pattern type"
    sheet['J2'] = "Pattern string"
    sheet['K2'] = "Pattern Behaviour"
    sheet['L2'] = "Replace string"
    sheet['M2'] = "On match"
    sheet['N2'] = "Target"
    sheet['O2'] = "State"

    for enum, keys in enumerate(dict_search_rule,3):
        sheet['A{0}'.format(enum)].value = dict_search_rule[keys]['Priority']
        sheet['B{0}'.format(enum)].value = dict_search_rule[keys]['Name']
        sheet['C{0}'.format(enum)].value = dict_search_rule[keys]['Description']
        sheet['D{0}'.format(enum)].value = dict_search_rule[keys]['Protocol']
        sheet['E{0}'.format(enum)].value = dict_search_rule[keys]['Source Mode']
        sheet['F{0}'.format(enum)].value = dict_search_rule.get(keys).get('Source Name',"")
        sheet['G{0}'.format(enum)].value = dict_search_rule[keys]['Authentication']
        sheet['H{0}'.format(enum)].value = dict_search_rule[keys]['Mode']
        sheet['I{0}'.format(enum)].value = dict_search_rule.get(keys).get('Pattern Type',"")
        sheet['J{0}'.format(enum)].value = dict_search_rule.get(keys).get('Pattern String',"")
        sheet['K{0}'.format(enum)].value = dict_search_rule.get(keys).get('Pattern Behavior',"")
        sheet['L{0}'.format(enum)].value = dict_search_rule.get(keys).get('Pattern Replace',"")
        sheet['M{0}'.format(enum)].value = dict_search_rule[keys]['Progress']
        sheet['N{0}'.format(enum)].value = dict_search_rule[keys]['Target Name']
        sheet['O{0}'.format(enum)].value = dict_search_rule[keys]['State']

    # 23) Configuration > Policy Service
    ##########################################################################
    ##########################################################################

    wb.create_sheet(title="Configuration > Policy Services")
    sheet = wb.get_sheet_by_name("Configuration > Policy Services")
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 20
    sheet.column_dimensions['I'].width = 20
    sheet.column_dimensions['J'].width = 20
    sheet.column_dimensions['K'].width = 20
    sheet.column_dimensions['L'].width = 20
    sheet.column_dimensions['M'].width = 20

    dict_policy_services = xconf_to_dict(xconfiguration_file,regex_policy_services)

    sheet.merge_cells('A1:M1')
    sheet['A1'] = "Policy services"
    sheet['A2'] = "Name"
    sheet['B2'] = "Description"
    sheet['C2'] = "Protocol"
    sheet['D2'] = "Certificate verification mode"
    sheet['E2'] = "HTTPS certificate revodation list (CRL) checking"
    sheet['F2'] = "Server 1 address"
    sheet['G2'] = "Server 2 address"
    sheet['H2'] = "Server 3 address"
    sheet['I2'] = "Path"
    sheet['J2'] = "Status path"
    sheet['K2'] = "Username"
    sheet['L2'] = "Password"
    sheet['M2'] = "Default CPL"

    if dict_policy_services:
        for enum, keys in enumerate(dict_policy_services,3):
            sheet['A{0}'.format(enum)].value = dict_policy_services[keys]['Name']
            sheet['B{0}'.format(enum)].value = dict_policy_services[keys]['Description']
            sheet['C{0}'.format(enum)].value = dict_policy_services[keys]['Protocol']
            sheet['D{0}'.format(enum)].value = dict_policy_services[keys]['TLS Verify Mode']
            sheet['E{0}'.format(enum)].value = dict_policy_services[keys]['TLS CRLCheck Mode']
            sheet['F{0}'.format(enum)].value = dict_policy_services[keys]['Server 1 Address']
            sheet['G{0}'.format(enum)].value = dict_policy_services[keys]['Server 2 Address']
            sheet['H{0}'.format(enum)].value = dict_policy_services[keys]['Server 3 Address']
            sheet['I{0}'.format(enum)].value = dict_policy_services[keys]['Path']
            sheet['J{0}'.format(enum)].value = dict_policy_services[keys]['Status Path']
            sheet['K{0}'.format(enum)].value = dict_policy_services[keys]['UserName']
            sheet['L{0}'.format(enum)].value = dict_policy_services[keys]['Password']
            sheet['M{0}'.format(enum)].value = dict_policy_services[keys]['DefaultCPL']

    # 24) Configuration > Bandwidth
    ##########################################################################
    ##########################################################################

    wb.create_sheet(title="Configuration > Bandwidth")
    sheet = wb.get_sheet_by_name("Configuration > Bandwidth")
    dict_bw = xconf_to_dict(xconfiguration_file,regex_bw)
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 20
    sheet.merge_cells('A1:B1')
    sheet['A1'] = "Bandwidth Configuration"
    sheet['A2'] = "Default Call bandwidth (kbps)"
    sheet['A3'] = "Downspeed per call mode"
    sheet['A4'] = "Downspeed total mode"
    sheet['B2'] = dict_bw['Bandwidth']['Default']
    sheet['B3'] = dict_bw['Bandwidth']['Downspeed PerCall Mode']
    sheet['B4'] = dict_bw['Bandwidth']['Downspeed Total Mode']

    # 25) Configuration > Links
    ##########################################################################
    ##########################################################################
    wb.create_sheet(title="Configuration > BW > Links")
    sheet = wb.get_sheet_by_name("Configuration > BW > Links")
    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 40
    sheet.column_dimensions['C'].width = 40
    sheet.column_dimensions['D'].width = 40
    sheet.column_dimensions['E'].width = 40

    dict_bw_link = xconf_to_dict(xconfiguration_file,regex_bw_link)

    sheet.merge_cells('A1:E1')
    sheet['A1'] = "Links"
    sheet['A2'] = "Name"
    sheet['B2'] = "Node 1"
    sheet['C2'] = "Node 2"
    sheet['D2'] = "Pipe 1"
    sheet['E2'] = "Pipe 2"

    for enum, keys in enumerate(sorted(dict_bw_link),3):
        sheet['A{0}'.format(enum)].value = dict_bw_link[keys]['Name']
        sheet['B{0}'.format(enum)].value = dict_bw_link[keys]['Node1 Name']
        sheet['C{0}'.format(enum)].value = dict_bw_link[keys]['Node2 Name']
        sheet['D{0}'.format(enum)].value = dict_bw_link[keys]['Pipe1 Name']
        sheet['E{0}'.format(enum)].value = dict_bw_link[keys]['Pipe2 Name']

    # 26) Configuration > Pipes
    ##########################################################################
    ##########################################################################
    wb.create_sheet(title="Configuration > BW > Pipes")
    sheet = wb.get_sheet_by_name("Configuration > BW > Pipes")
    sheet.column_dimensions['A'].width = 40

    dict_bw_pipe =xconf_to_dict(xconfiguration_file,regex_bw_pipe)

    sheet.merge_cells('A1:D1')
    sheet['A1'].style = 'styleObj'
    sheet['A1'] = "Configuration"
    sheet['A2'] = "Name"
    sheet.merge_cells('A3:D3')
    sheet['A3'].style = 'styleObj'
    sheet['A3'] = "Total bandwidth available"
    sheet['A4'] = "Bandwidth restriction"
    sheet['A5'] = "Total bandwidth limit (kbps)"
    sheet.merge_cells('A6:D6')
    sheet['A6'].style = 'styleObj'
    sheet['A6'] = "Calls through this pipe"
    sheet['A7'] = "Bandwidth restriction"
    sheet['A8'] = "Total bandwidth limit (kbps)"

    if dict_bw_pipe:
        for enum, keys in enumerate(sorted(dict_bw_pipe),2):
            sheet.cell(row = 2, column = enum).value = dict_bw_pipe[keys]['Name']
            sheet.cell(row = 4, column = enum).value = dict_bw_pipe[keys]['Bandwidth Total Mode']
            sheet.cell(row = 5, column = enum).value = dict_bw_pipe.get(keys).get('Bandwidth Total Limit',"")
            sheet.cell(row = 7, column = enum).value = dict_bw_pipe[keys]['Bandwidth PerCall Mode']
            sheet.cell(row = 8, column = enum).value = dict_bw_pipe.get(keys).get('Bandwidth PerCall Limit',"")


    ##########################################################################
    ##########################################################################
    ###################  VCS Expressway only information #####################
    ##########################################################################
    ##########################################################################
    dict_traversal_ports = xconf_to_dict(xconfiguration_file,regex_traversal_ports)
    dict_turn = xconf_to_dict(xconfiguration_file,regex_turn)
    dict_traversal_endpoints = xconf_to_dict(xconfiguration_file,regex_traversal_endpoints)

    try:
        if dict_traversal_ports['Server Media']['Demultiplexing UseConfiguredDemuxPorts']:
        # Configuration > Traversal > Ports
            wb.create_sheet(title="Traversal > Ports")
            sheet = wb.get_sheet_by_name("Traversal > Ports")
            sheet.column_dimensions['A'].width = 40
            sheet.column_dimensions['B'].width = 20
            sheet.merge_cells('A1:B1')
            sheet['A1'].style = 'styleObj'
            sheet['A1'] = "Demultiplexing ports"
            sheet['A2'] = "Use configured demultiplexing ports"
            sheet['A3'] = "Media demultiplexing RTP port"
            sheet['A4'] = "Media demultiplexing RTCP port"
            sheet.merge_cells('A5:B5')
            sheet['A5'].style = 'styleObj'
            sheet['A5'] = "Call signalling ports"
            sheet['A6'] = "H.323 Assent call signaling port"
            sheet['A7'] = "H.323 H.460.18 call signaling port"
            sheet['B2'] = dict_traversal_ports['Server Media']['Demultiplexing UseConfiguredDemuxPorts']
            sheet['B3'] = dict_traversal_ports['Server Media']['Demultiplexing RTP Port']
            sheet['B4'] = dict_traversal_ports['Server Media']['Demultiplexing RTCP Port']
            sheet['B6'] = dict_traversal_ports['Server H323']['Assent CallSignaling Port']
            sheet['B7'] = dict_traversal_ports['Server H323']['H46018 CallSignaling Port']
    except:
        logging.debug("This is not a VCS Expresway or Expressway-E")

    # Configuration > Traversal > TURN
    if dict_turn:
        wb.create_sheet(title="Traversal > TURN")
        sheet = wb.get_sheet_by_name("Traversal > TURN")
        sheet.column_dimensions['A'].width = 40
        sheet.column_dimensions['B'].width = 20
        sheet.merge_cells('A1:B1')
        sheet['A1'].style = 'styleObj'
        sheet['A1'] = "Server"
        sheet['A2'] = "TURN services"
        sheet['A3'] = "TURN requests port"
        sheet['A4'] = "Delegated credential checing"
        sheet['A5'] = "Authentication realm"
        sheet['A6'] = "Media port range start"
        sheet['A7'] = "Media port range end"
        sheet['B2'] = dict_turn['TURN']['Mode']
        sheet['B3'] = dict_turn['TURN']['PortRangeStart']
        sheet['B4'] = dict_turn['TURN']['Authentication Remote Mode']
        sheet['B5'] = dict_turn['TURN']['Authentication Realm']
        sheet['B6'] = dict_turn['TURN']['Media Port Start']
        sheet['B7'] = dict_turn['TURN']['Media Port End']

    # Configuration > Traversal > Locally registered endpoints
    if dict_traversal_endpoints:
        wb.create_sheet(title="Traversal > Loc reg endp")
        sheet = wb.get_sheet_by_name("Traversal > Loc reg endp")
        sheet.column_dimensions['A'].width = 40
        sheet.column_dimensions['B'].width = 20
        sheet.merge_cells('A1:B1')
        sheet['A1'].style = 'styleObj'
        sheet['A1'] = "Configuration"
        sheet['A2'] = "H.323 Assent mode"
        sheet['A3'] = "H.460.18 mode"
        sheet['A4'] = "H.460.19 demultiplexing mode"
        sheet['A5'] = "H.323 preference"
        sheet['A6'] = "UDP probe retry interval"
        sheet['A7'] = "UDP probe retry count"
        sheet['A8'] = "UDP probe keep alive interval"
        sheet['A9'] = "TCP probe retry interval"
        sheet['A10'] = "TCP probe retry count"
        sheet['A11'] = "TCP probe keep alive interval"
        sheet['B2'] = dict_traversal_endpoints['Assent']['Mode']
        sheet['B3'] = dict_traversal_endpoints['H46018']['Mode']
        sheet['B4'] = dict_traversal_endpoints['H46019 Demultiplexing']['Mode']
        sheet['B5'] = xconf_to_dict(xconfiguration_file,regex_h323_pref)['H323']['Preference']
        sheet['B6'] = dict_traversal_endpoints['UDPProbe']['RetryInterval']
        sheet['B7'] = dict_traversal_endpoints['UDPProbe']['RetryCount']
        sheet['B8'] = dict_traversal_endpoints['UDPProbe']['KeepAliveInterval']
        sheet['B9'] = dict_traversal_endpoints['TCPProbe']['RetryInterval']
        sheet['B10'] = dict_traversal_endpoints['TCPProbe']['RetryCount']
        sheet['B11'] = dict_traversal_endpoints['TCPProbe']['RetryInterval']


    # 27) Application > Conference Factory
    ##########################################################################
    ##########################################################################
    dict_multiway = xconf_to_dict(xconfiguration_file,regex_multiway)

    if dict_multiway:
        wb.create_sheet(title="App > Conference Factory")
        sheet = wb.get_sheet_by_name("App > Conference Factory")
        sheet.column_dimensions['A'].width = 40
        sheet.column_dimensions['B'].width = 20

        sheet.merge_cells('A1:B1')
        sheet['A1'].style = 'styleObj'
        sheet['A1'] = "Configuration"
        sheet['A2'] = "Mode"
        sheet['A3'] = "Alias"
        sheet['A4'] = "Template"
        sheet['A5'] = "Number range start"
        sheet['A6'] = "Number range end"

        sheet['B2'] = dict_multiway['ConferenceFactory']['Mode']
        sheet['B3'] = dict_multiway['ConferenceFactory']['Alias']
        sheet['B4'] = dict_multiway['ConferenceFactory']['Template']
        sheet['B5'] = dict_multiway['ConferenceFactory']['Range Start']
        sheet['B6'] = dict_multiway['ConferenceFactory']['Range End']

    # 28) Application > Presence
    ##########################################################################
    ##########################################################################
    wb.create_sheet(title="App > Presence")
    sheet = wb.get_sheet_by_name("App > Presence")
    dict_presence = xconf_to_dict(xconfiguration_file,regex_presence)

    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 20
    sheet.merge_cells('A1:B1')
    sheet['A1'].style = 'styleObj'
    sheet['A1'] = "PUA"
    sheet['A2'] = "SIP SIMPLE Presence User Agent"
    sheet['A3'] = "Default published status for registered endpoints"
    sheet.merge_cells('A4:B4')
    sheet['A4'].style = 'styleObj'
    sheet['A4'] = "Presence Server"
    sheet['A5'] = "SIP SIMPLE Presence Server"
    sheet['B2'] = dict_presence['User']['Agent Mode']
    sheet['B3'] = dict_presence['User']['Agent Presentity Idle Status']
    sheet['B5'] = dict_presence['Server']['Mode']

    # 29) Application > FindMe
    ##########################################################################
    ##########################################################################
    wb.create_sheet(title="App > FindMe")
    sheet = wb.get_sheet_by_name("App > FindMe")
    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 20
    dict_findme = xconf_to_dict(xconfiguration_file,regex_findme)

    sheet.merge_cells('A1:B1')
    sheet['A1'].style = 'styleObj'
    sheet['A1'] = "Configuration"
    sheet['A2'] = "FindMe mode"
    findme_mode = dict_findme['FindMe']['Mode']

    if findme_mode == "Off":
        sheet['B2'] = findme_mode
    elif findme_mode == "On":
        sheet['B2'] = findme_mode
        sheet['A3'] = "Caller ID"
        sheet['A4'] = "Cluster name (FQDN for provisioning)"
        sheet['B3'] = dict_findme['FindMe']['CallerId']
        sheet['B4'] = dict_cluster['Alternates']['Cluster Name']
    else:
        sheet['B2'] = findme_mode
        sheet['A3'] = "Caller ID"
        sheet['A4'] = "Protocol"
        sheet['A5'] = "Address"
        sheet['A6'] = "Path"
        sheet['A7'] = "Username"
        sheet['A8'] = "Password"
        sheet['A9'] = "Cluster name (FQDN for provisioning)"

        sheet['B3'] = dict_findme['FindMe']['CallerId']
        sheet['B4'] = dict_findme['FindMe']['Server Protocol']
        sheet['B5'] = dict_findme['FindMe']['Server Address']
        sheet['B6'] = dict_findme['FindMe']['Server Path']
        sheet['B7'] = dict_findme['FindMe']['Server UserName']
        sheet['B8'] = dict_findme['FindMe']['Server Password']
        sheet['B9'] = dict_cluster['Alternates']['Cluster Name']

    # 30) Users > Password security
    ##########################################################################
    ##########################################################################
    wb.create_sheet(title="Users > Password Sec")
    sheet = wb.get_sheet_by_name("Users > Password Sec")
    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 20
    sheet.merge_cells('A1:B1')
    sheet['A1'].style = 'styleObj'
    sheet['A1'] = "Password security"
    sheet['A2'] = "Enforce strict passwords"
    try:
        sheet['B2'] = dict_password_security['StrictPassword']['Enabled',"Not defined"]
    except KeyError:
        sheet['B2'] = "Not defined"


    # 31) Users > Administrator accounts
    ##########################################################################
    ##########################################################################
    wb.create_sheet(title="Users > Admin Accounts")
    sheet = wb.get_sheet_by_name("Users > Admin Accounts")
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20

    sheet.merge_cells('A1:E1')
    sheet['A1'].style = 'styleObj'
    sheet['A1'] = "Administrator Accounts"
    sheet['A2'] = "Name"
    sheet['B2'] = "State"
    sheet['C2'] = "Access level"
    sheet['D2'] = "Web access"
    sheet['E2'] = "API access"

    for enum, keys in enumerate(sorted(dict_user_admin), 3):
        sheet['A{0}'.format(enum)] = keys
        sheet['B{0}'.format(enum)] = dict_user_admin[keys]['Enabled']
        sheet['D{0}'.format(enum)] = dict_user_admin[keys]['AccessWeb']
        sheet['E{0}'.format(enum)] = dict_user_admin[keys]['AccessAPI']


    # 32) Users > Administrator groups
    ##########################################################################
    ##########################################################################
    wb.create_sheet(title="Users > Admin groups")
    sheet = wb.get_sheet_by_name("Users > Admin groups")
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20

    sheet.merge_cells('A1:E1')
    sheet['A1'].style = 'styleObj'
    sheet['A1'] = "Administrator Groups"
    sheet['A2'] = "Name"
    sheet['B2'] = "State"
    sheet['C2'] = "Access level"
    sheet['D2'] = "Web access"
    sheet['E2'] = "API access"

    for enum, keys in enumerate(dict_group_admin,3):
        sheet['A{0}'.format(enum)] = keys
        sheet['B{0}'.format(enum)] = dict_group_admin[keys]['Enabled']
        sheet['D{0}'.format(enum)] = dict_group_admin[keys]['AccessWeb']
        sheet['E{0}'.format(enum)] = dict_group_admin[keys]['AccessAPI']


    # 33) Users > LDAP Configuration
    ##########################################################################
    ##########################################################################
    wb.create_sheet(title="Users > LDAP Configuration")
    sheet = wb.get_sheet_by_name("Users > LDAP Configuration")
    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20

    sheet.merge_cells('A1:C1')
    sheet['A1'].style = 'styleObj'
    sheet['A1'] = "Remote authentication source"
    sheet['A2'] = "Administrator authentication source"
    try:
        sheet['B2'] = dict_login_source['Login Source']['Admin']
    except KeyError:
        sheet['B2'] = "N/A to this version"
    sheet['A3'] = "FindMe authentication source"
    try:
        sheet['B3'] = dict_login_source['Login Source']['User']
    except KeyError:
        sheet['B3']= "N/A to this version"

    try:
        if (dict_login_source['Login Source']['Admin'] == "Remote"
            or dict_login_source['Login Source']['Admin'] == "Both"):
            sheet.merge_cells('A4:C4')
            sheet['A4'].style = 'styleObj'
            sheet['A4'] = "LDAP server configuration"
            fqdn_resolution = dict_remote_login['Login Remote']['LDAP Server FQDNResolution']
            sheet['A5'] = "FQDN address resolution"
            sheet['B5'] = fqdn_resolution

            if fqdn_resolution == "IPAddress":
                sheet['A6'] = "Server address"
                sheet['B6'] = dict_remote_login['Login Remote']['LDAP Server Address']
                sheet['A7'] = "Port"
                sheet['B7'] = dict_remote_login['Login Remote']['LDAP Server Port']
            elif fqdn_resolution == "AddressRecord" or fqdn_resolution == "SRVRecord":
                sheet['A6'] = "Host name and Domain"
                ldap_server_address = dict_remote_login['Login Remote']['LDAP Server Address']

            if fqdn_resolution == "AddressRecord":
                sheet['B6'] = ldap_server_address.split(".")[0]
                sheet['A7'] = "Port"
                sheet['B7'] = dict_remote_login['Login Remote']['LDAP Server Port']
                try:
                    sheet['C6'] = ldap_server_address.split(".")[1]
                except:
                    sheet['C6'] = ""
            else:
                sheet['B6'] = ""
                sheet['C6'] = ldap_server_address
                sheet['A7'] = "Port"
                sheet['B7'] = ""
        sheet['A8'] = "Encryption"
        sheet['B8'] = dict_remote_login['Login Remote']['LDAP Encryption']
        sheet['A9'] = "Certificate revocation list (CRL) checking"
        sheet['B9'] = dict_remote_login['Login Remote']['LDAP CRLCheck']
        sheet.merge_cells('A10:C10')
        sheet['A10'].style = 'styleObj'
        sheet['A10'] = "Authentication configuration"
        sheet['A11'] = "Bind DN"
        sheet['A12'] = "Bind password"
        sheet['A13'] = "Bind SASL"
        sheet['A14'] = "Bind username"
        sheet['B11'] = dict_remote_login['Login Remote']['LDAP VCS BindDN']
        sheet['B12'] = dict_remote_login['Login Remote']['LDAP VCS BindPassword']
        sheet['B13'] = dict_remote_login['Login Remote']['LDAP SASL']
        sheet['B14'] = dict_remote_login['Login Remote']['LDAP VCS BindUsername']
        sheet.merge_cells('A15:C15')
        sheet['A15'].style = 'styleObj'
        sheet['A15'] = "Directory configuration"
        sheet['A16'] = "Base DN for accounts"
        sheet['A17'] = "Base DN for groups"
        sheet['B16'] = dict_remote_login['Login Remote']['LDAP BaseDN Accounts']
        sheet['B17'] = dict_remote_login['Login Remote']['LDAP BaseDN Groups']

    except:
        sheet.merge_cells('A4:C4')
        sheet['A4'].style = 'styleObj'
        sheet['A4'] = "LDAP server configuration"
        fqdn_resolution = dict_remote_login['Login Remote']['LDAP Server FQDNResolution']

        sheet['A5'] = "FQDN address resolution"
        sheet['B5'] = fqdn_resolution

        if fqdn_resolution == "IPAddress":
            sheet['A6'] = "Server address"
            sheet['B6'] = dict_remote_login['Login Remote']['LDAP Server Address']
            sheet['A7'] = "Port"
            sheet['B7'] = dict_remote_login['Login Remote']['LDAP Server Port']
        elif fqdn_resolution == "AddressRecord" or fqdn_resolution == "SRVRecord":
            sheet['A6'] = "Host name and Domain"
            ldap_server_address = dict_remote_login['Login Remote']['LDAP Server Address']

        if fqdn_resolution == "AddressRecord":
            sheet['B6'] = ldap_server_address.split(".")[0]
            sheet['A7'] = "Port"
            sheet['B7'] = dict_remote_login['Login Remote']['LDAP Server Port']
            try:
                sheet['C6'] = ldap_server_address.split(".")[1]
            except:
                sheet['C6'] = ""
        else:
            sheet['B6'] = ""
            sheet['C6'] = ldap_server_address
            sheet['A7'] = "Port"
            sheet['B7'] = ""
        sheet['A8'] = "Encryption"
        sheet['B8'] = dict_remote_login['Login Remote']['LDAP Encryption']
        sheet['A9'] = "Certificate revocation list (CRL) checking"
        sheet['B9'] = dict_remote_login['Login Remote']['LDAP CRLCheck']
        sheet.merge_cells('A10:C10')
        sheet['A10'].style = 'styleObj'
        sheet['A10'] = "Authentication configuration"
        sheet['A11'] = "Bind DN"
        sheet['A12'] = "Bind password"
        sheet['A13'] = "Bind SASL"
        sheet['A14'] = "Bind username"
        sheet['B11'] = dict_remote_login['Login Remote']['LDAP VCS BindDN']
        sheet['B12'] = dict_remote_login['Login Remote']['LDAP VCS BindPassword']
        sheet['B13'] = dict_remote_login['Login Remote']['LDAP SASL']
        sheet['B14'] = dict_remote_login['Login Remote']['LDAP VCS BindUsername']
        sheet.merge_cells('A15:C15')
        sheet['A15'].style = 'styleObj'
        sheet['A15'] = "Directory configuration"
        sheet['A16'] = "Base DN for accounts"
        sheet['A17'] = "Base DN for groups"
        sheet['B16'] = dict_remote_login['Login Remote']['LDAP BaseDN Accounts']
        sheet['B17'] = dict_remote_login['Login Remote']['LDAP BaseDN Groups']

    # 34) Maintenance > Maintenance mode
    ##########################################################################
    ##########################################################################
    wb.create_sheet(title="Maintenance mode")
    sheet = wb.get_sheet_by_name("Maintenance mode")
    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 20

    sheet.merge_cells('A1:B1')
    sheet['A1'].style = 'styleObj'
    sheet['A1'] = "Configuration"
    sheet['A2'] = "Maintenance mode"
    sheet['B2'] = xconf_to_dict(xconfiguration_file,regex_maintenance)['Maintenance']['Mode']


    # Save workbook using as file name the VCS system name and the time the program was initiated.
    wb.save(destination_filename)


if __name__=="__main__":
    xconf2excel(sys.argv[1])
